from flask import request
from flask_restful.reqparse import RequestParser
from flask_restful import Resource, marshal_with, fields
from app.api_0_1.common.utils import make_response, make_marshal_fields, convert_pagination, \
    make_response_from_pagination
from app.models import db
from app.models.common.utils import paginate
from app.models.entertainment_city import EntertainmentCity
from app.models.member import Member
from app.models.config_yongjin import ConfigYongjin, YongJinTyb
import json
from sqlalchemy import and_
from app.api_0_1.common import DEFAULT_PAGE, DEFAULT_PAGE_SIZE
from app.models.member_yongjin_compute import MemberYongjinCompute
from sqlalchemy import func
from flask import current_app

'''
系统设置 - 佣金
'''
# 返回娱乐城和对应的游戏类型
'''		AG:{"1002":0.9,"1003":0.8,"1003":0.7},
		BB:{"1002":0.4,"1003":0.3}'''


class YlcAndGametype(Resource):

    def get(self):
        data = {}
        query = db.session.query(
            EntertainmentCity.code,
            EntertainmentCity.game_types).all()
        ecmap = {}
        for i in query:
            eccode = i.code
            ecmap[eccode] = {}
            gtyps = i.game_types
            if gtyps is None or gtyps == "":
                continue
            typeList = gtyps.split(",")
            for type in typeList:
                ecmap[eccode][type] = None
        data["tuiyongbi"] = ecmap
        data['pcJine'] = None
        data['yxhuiyuan'] = None
        data['youhui'] = None
        data['fanshui'] = None
        return make_response(data)


# 佣金设定列表
class ConfigYongjinAPI(Resource):
    @marshal_with(make_marshal_fields({
        'id': fields.Integer,
        'name': fields.String,
        'enable': fields.Integer,
        'zdtze': fields.Float,
        'cksxfsx': fields.Float,
        'qksxfsx': fields.Float,
        'dbcksxf': fields.Float,
        'dbqksxf': fields.Float,
        'zdckje': fields.Float,
        'dls': fields.Integer}))
    def get(self):
        query = db.session.query(
            ConfigYongjin.id,
            ConfigYongjin.name,
            ConfigYongjin.enable,
            ConfigYongjin.zdtze,
            ConfigYongjin.dbcksxf,
            ConfigYongjin.dbqksxf,
            ConfigYongjin.cksxfsx,
            ConfigYongjin.qksxfsx,
            ConfigYongjin.zdckje)
        pagination = paginate(query)
        query1 = db.session.query(ConfigYongjin.id).all()
        res = []
        result = []
        for id in query1:
            a = id[0]
            count = db.session.query(Member).filter(and_(Member.commissionConfig == a, Member.type != 0)).count()
            res.append(count)
        for i in range(len(pagination.items)):
            item = dict(zip(pagination.items[i].keys(), pagination.items[i]))
            item['dls'] = res[i]
            result.append(item)
        return make_response(result)

    # 	新增佣金设定
    def post(self):
        data = request.get_json()
        data = data['data']
        configyongjin = ConfigYongjin()
        try:
            if 'enable' in data:
                if data['enable'] is not '':
                    configyongjin.enable = data['enable']
            if 'name' in data:
                if data['name'] is not '':
                    configyongjin.name = data['name']
                else:
                    return {"errorMsg": "标题不能为空"}
            if 'zdtze' in data:
                if data['zdtze'] is not '':
                    configyongjin.zdtze = data['zdtze']
                else:
                    return {"errorMsg": "最低投注额不能为空"}
            if 'dbcksxf' in data:
                if data['dbcksxf'] is not '':
                    configyongjin.dbcksxf = data['dbcksxf']
                else:
                    return {"errorMsg": "单笔存款手续费不能为空"}


            if 'dbqksxf' in data:
                if data['dbqksxf'] is not '':
                    configyongjin.dbqksxf = data['dbqksxf']
                else:
                    return {"errorMsg": "单笔取款手续费不能为空"}

            if 'cksxfsx' in data:
                if data['cksxfsx'] is '':
                    configyongjin.cksxfsx = None
                else:
                    configyongjin.cksxfsx = data['cksxfsx']

            if 'qksxfsx' in data:
                if data['qksxfsx'] is '':
                    configyongjin.qksxfsx = None
                else:
                    configyongjin.qksxfsx = data['qksxfsx']

            if 'zdckje' in data:
                if data['zdckje'] is '':
                    configyongjin.zdckje = None
                else:
                    configyongjin.zdckje = data['zdckje']


            if 'fendang' in data:
                fendang = data['fendang']
                if not fendang:
                    return {"errorMsg": "派彩,有效会员,退佣比,优惠,返水是必填项"}

                db.session.add(configyongjin)
                # db.session.commit()
                db.session.flush()

                for i in fendang:
                    tyb = YongJinTyb()
                    tyb.Yid = configyongjin.id
                    if i['pcJine'] is not None:
                        tyb.pcJine = i['pcJine']
                    else:
                        raise  Exception("派彩金额不能为空")

                    if i['yxhuiyuan'] is not None:
                        tyb.yxhuiyuan = i['yxhuiyuan']
                    else:
                        raise  Exception("有效会员不能为空")

                    if i['youhui'] is not None:
                        tyb.youhui = i['youhui']
                    else:
                        raise  Exception("优惠不能为空")

                    if i['fanshui'] is not None:
                        tyb.fanshui = i['fanshui']
                    else:
                        raise Exception("返水不能为空")

                    if i['tuiyongbi'] is not None:
                        j = json.dumps(i['tuiyongbi'])
                        tyb.tuiyongbi = j
                    db.session.add(tyb)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            db.session.remove()
            # if configyongjin.id:
            #     db.session.delete(configyongjin)
            #     try:
            #         db.session.commit()
            #     except:
            #         return {'success': False, 'errorMsg': '添加失败'}

            return {'success': False, "errorMsg":'派彩,有效会员,退佣比,优惠,返水是必填项'}
        return {'success': True, 'errorMsg': '添加成功'}


# 获取详情页
class YongJinContent(Resource):

    @marshal_with(make_marshal_fields({
        'id': fields.Integer,
        'enable': fields.Integer,
        'name': fields.String,
        'zdtze': fields.Float,
        'dbcksxf': fields.Float,
        'cksxfsx': fields.Float,
        'dbqksxf': fields.Float,
        'qksxfsx': fields.Float,
        'zdckje': fields.Float,
        'dls': fields.Integer,
        'fendang': fields.String

    }))
    def get(self, id):
        data = {}
        count = db.session.query(Member).filter(and_(Member.commissionConfig == id, Member.type != 0)).count()
        query = db.session.query(
            ConfigYongjin.id,
            ConfigYongjin.cksxfsx,
            ConfigYongjin.qksxfsx,
            ConfigYongjin.enable,
            ConfigYongjin.zdtze,
            ConfigYongjin.dbcksxf,
            ConfigYongjin.dbqksxf,
            ConfigYongjin.name,
            ConfigYongjin.zdckje).filter(ConfigYongjin.id == id)
        for a in query:
            data['id'] = id
            data['enable'] = a.enable
            data['zdtze'] = a.zdtze
            data['dbqksxf'] = a.dbqksxf
            data['dbcksxf'] = a.dbcksxf
            data['zdckje'] = a.zdckje
            data['dls'] = count
            data['cksxfsx'] = a.cksxfsx
            data['qksxfsx'] = a.qksxfsx
            data['name'] = a.name
        query1 = db.session.query(
            YongJinTyb.id,
            YongJinTyb.pcJine,
            YongJinTyb.youhui,
            YongJinTyb.yxhuiyuan,
            YongJinTyb.fanshui,
            YongJinTyb.tuiyongbi).filter(YongJinTyb.Yid == id).all()
        data['fendang'] = []
        for i in query1:
            a = {}
            a['id'] = i[0]
            a['pcJine'] = i[1]
            a['youhui'] = i[2]
            a['yxhuiyuan'] = i[3]
            a['fanshui'] = i[4]
            a['tuiyongbi'] = i[5]
            data['fendang'].append(a)
        data['fendang'] = json.dumps((data['fendang']))

        return make_response(data)

    def put(self, id):
        data = request.get_json()
        data = data['data']
        configyongjin = ConfigYongjin.query.get(id)
        try:
            if data['name']:
                configyongjin.name = data['name']
            if data['zdtze']:
                configyongjin.zdtze = data['zdtze']
            if data['dbcksxf']:
                configyongjin.dbcksxf = data['dbcksxf']
            if data['cksxfsx']:
                configyongjin.cksxfsx = data['cksxfsx']
            if data['dbqksxf']:
                configyongjin.dbqksxf = data['dbqksxf']
            if data['qksxfsx']:
                configyongjin.qksxfsx = data['qksxfsx']
            if data['zdckje']:
                configyongjin.zdckje = data['zdckje']
            try:
                db.session.add(configyongjin)
            except:
                db.session.rollback()
                db.session.remove()

            if 'delete' in data:
                num = data['delete']
                try:
                    for dd in num:
                        yjtyb = YongJinTyb.query.filter(YongJinTyb.id == dd).first()
                        db.session.delete(yjtyb)
                except:
                    db.session.rollback()
                    db.session.remove()

            new = data['fendang']
            for i in new:
                if 'id' in i:
                    id = i['id']
                    yjtyb = YongJinTyb.query.get(id)
                    if i['pcJine']:
                        yjtyb.pcJine = i['pcJine']
                    if i['yxhuiyuan']:
                        yjtyb.yxhuiyuan = i['yxhuiyuan']
                    if i['pcJine']:
                        yjtyb.youhui = i['youhui']
                    if i['fanshui']:
                        yjtyb.fanshui = i['fanshui']
                    if i['tuiyongbi']:
                        yjtyb.tuiyongbi = json.dumps(i['tuiyongbi'])
                    try:
                        db.session.add(yjtyb)
                    except:
                        db.session.rollback()
                        db.session.remove()

                if "id" not in i:
                    try:
                        tyb = YongJinTyb()
                        tyb.Yid = configyongjin.id
                        if i['pcJine'] is not None:
                            tyb.pcJine = i['pcJine']
                        else:
                            return {"errorMsg": "派彩金额不能为空"}
                        if i['yxhuiyuan'] is not None:
                            tyb.yxhuiyuan = i['yxhuiyuan']
                        else:
                            return {"errorMsg": "有效会员不能为空"}
                        if i['youhui'] is not None:
                            tyb.youhui = i['youhui']
                        else:
                            return {"errorMsg": "优惠不能为空"}
                        if i['fanshui'] is not None:
                            tyb.fanshui = i['fanshui']
                        else:
                            return {"errorMsg": "返水不能为空"}
                        if i['tuiyongbi'] is not None:
                            j = json.dumps(i['tuiyongbi'])
                            tyb.tuiyongbi = j

                        db.session.add(tyb)
                    except:
                        db.session.rollback()
                        db.session.remove()
            db.session.commit()
        except:
            db.session.rollback()
            db.session.remove()
        return {'success': True, 'messages': '修改成功'}

    def delete(self, id):
        count = db.session.query(Member).filter(and_(Member.commissionConfig == id, Member.type != 0)).count()
        if count >= 1:
            return {'success': False, 'errorMsg': '此设定有代理不可删除'}
        else:
            enen = db.session.query(ConfigYongjin).get(id)
            db.session.delete(enen)
            db.session.commit()
            tyb = db.session.query(YongJinTyb).filter(YongJinTyb.Yid == id).first()
            if tyb:
                db.session.delete(tyb)
                db.session.commit()
            return {'success': True}


class ChangeStatus(Resource):
    def put(self):
        parser = RequestParser()
        parser.add_argument('id', type=int)
        parser.add_argument('enable', type=int)
        args = parser.parse_args(strict=True)
        args = {key: value for key, value in args.items() if value is not None}

        try:
            ConfigYongjin.query.filter(ConfigYongjin.id == args["id"]).update(args)
            db.session.commit()
        except:
            db.session.rollback()
            db.session.remove()
        return {'success': True, 'messages': '修改成功'}


'''佣金计算'''
import time
import datetime
import copy
from app.models.member_fanshui_pc import MemberFanshuiPc
from app.models.blast_bets import BlastBets, BlastBetsCredit
from app.models.member_account_change import MemberAccountChangeRecord, Deposit, Withdrawal
from app.models.entertainment_city_bets_detail import EntertainmentCityBetsDetail
from sqlalchemy import and_, or_
from app.models.member_yongjin_compute import MemberYongjinCompute, MemberAgentDetail, MemberAgentExport
from sqlalchemy import func
from flask import current_app
import decimal


class DecimalEncoder(json.JSONEncoder):
    def default(self, o):
        if isinstance(o, decimal.Decimal):
            return float(o)
        super(DecimalEncoder, self).default(o)


'''代理佣金计算'''


class AgentYongjinCompute(Resource):
    def get(self):
        parser = RequestParser(trim=True)
        parser.add_argument('startTime', type=str)
        parser.add_argument('endTime', type=str)
        args = parser.parse_args(strict=True)

        startTime = request.args.get('startTime')
        endTime = request.args.get('endTime')
        # 最后一天, 时间+1
        date_list = time.strptime(endTime, "%Y-%m-%d")
        y, m, d = date_list[:3]
        delta = datetime.timedelta(days=1)
        date_result = datetime.datetime(y, m, d) + delta
        end_date = date_result.strftime("%Y-%m-%d")

        start_time = int(time.mktime(time.strptime(startTime, '%Y-%m-%d')))
        end_time = int(time.mktime(time.strptime(end_date, '%Y-%m-%d')))

        try:
            res_recordId = db.session.query(MemberYongjinCompute.recordId).filter().order_by(
                MemberYongjinCompute.recordId.desc()).first()

            recordId = res_recordId[0] + 1
        except:
            db.session.rollback()
            recordId = 1

        # 取出每个用户每天的投注金额
        result_bets_betamount = db.session.query(
            BlastBets.uid,
            BlastBets.username,
            func.sum(BlastBets.mode * BlastBets.beiShu * BlastBets.actionNum).label("betAmount"),
        ).filter(BlastBets.state == 2, BlastBets.actionTime.between(start_time, end_time)) \
            .group_by(BlastBets.uid, BlastBets.username).all()

        result_credit_betamount = db.session.query(
            BlastBetsCredit.memberId,
            BlastBetsCredit.memberUsername,
            func.sum(BlastBetsCredit.betAmount)
        ).filter(BlastBetsCredit.state == 2, BlastBetsCredit.betTime.between(start_time, end_time)) \
            .group_by(BlastBetsCredit.memberId, BlastBetsCredit.memberUsername).all()

        # 将聚合查询得到的decimal数据转为float, 再转回Python对象
        data_bets = json.dumps(result_bets_betamount, cls=DecimalEncoder)
        data_bets = json.loads(data_bets)
        data_credit = json.dumps(result_credit_betamount, cls=DecimalEncoder)
        data_credit = json.loads(data_credit)

        dict1 = {a[0]: a for a in data_bets}
        for b in data_credit:
            if b[0] in dict1:
                dict1[b[0]][2] += b[2]
            else:
                dict1[b[0]] = b
        data = list(dict1.values())
        data_betAmount = copy.deepcopy(data)

        # 　取出每个用户每天的总中奖金额
        result_bets_bonus = db.session.query(
            BlastBets.uid,
            BlastBets.username,
            func.sum(BlastBets.bonus),
        ).filter(BlastBets.state == 2, BlastBets.actionTime.between(start_time, end_time)) \
            .group_by(BlastBets.uid, BlastBets.username).all()

        result_credit_bonus = db.session.query(
            BlastBetsCredit.memberId,
            BlastBetsCredit.memberUsername,
            func.sum(BlastBetsCredit.bonus)
        ).filter(BlastBetsCredit.state == 2, BlastBetsCredit.betTime.between(start_time, end_time)) \
            .group_by(BlastBetsCredit.memberId, BlastBetsCredit.memberUsername).all()

        # 将聚合查询得到的decimal数据转为float, 再转回Python对象
        data_bets_bonus = json.dumps(result_bets_bonus, cls=DecimalEncoder)
        data_bets_bonus = json.loads(data_bets_bonus)
        data_credit_bonus = json.dumps(result_credit_bonus, cls=DecimalEncoder)
        data_credit_bonus = json.loads(data_credit_bonus)

        dict2 = {a[0]: a for a in data_bets_bonus}
        for b in data_credit_bonus:
            if b[0] in dict2:
                dict2[b[0]][2] += b[2]
            else:
                dict2[b[0]] = b
        data_bonus = list(dict2.values())

        # KK的损益
        dict3 = {a[0]: a for a in data}
        for b in data_bonus:
            if b[0] in dict3:
                dict3[b[0]][2] -= b[2]
            else:
                dict3[b[0]] = b
        data_sunyi = list(dict3.values())
        data_sunyi_copy = copy.deepcopy(data_sunyi)

        # 　再加入娱乐城的损益
        result_EC_betAmount = db.session.query(
            EntertainmentCityBetsDetail.PlayerName,
            func.sum(EntertainmentCityBetsDetail.Profit)
        ).filter(EntertainmentCityBetsDetail.BetTime.between(start_time, end_time)) \
            .group_by(EntertainmentCityBetsDetail.PlayerName).all()

        result_EC_betAmount_list = []
        for res in result_EC_betAmount:
            result_EC_list = []
            reb = db.session.query(Member.id).filter(Member.username == res[0]).first()
            result_EC_list.append(reb[0])
            result_EC_list.append(res[0])
            result_EC_list.append(res[1])
            result_EC_betAmount_list.append(result_EC_list)

        dict_4 = {a[0]: a for a in data_sunyi}
        for b in result_EC_betAmount_list:
            if b[0] in dict_4:
                dict_4[b[0]][2] -= b[2]
            else:
                dict_4[b[0]] = b
        data_total_sunyi = list(dict_4.values())

        # 把每个代理时间区间内的总损益和总有效会员人数查出来
        agent_list = set()  # set集合存代理id,  去重
        for i in data_total_sunyi:
            try:
                reb = db.session.query(Member.parent).filter(Member.id == i[0], Member.isTsetPLay != 1).first()
                agent_list.add(reb[0])
            except:
                db.session.rollback()
                continue

        list_a = []  # list_a每个代理的id、每个代理对应的会员的损益总和、每个代理对应的会员人数
        for i_id in agent_list:
            list_b = []
            sunyi = 0
            n = 0
            for i in data_total_sunyi:
                try:
                    reb = db.session.query(Member.parent).filter(Member.id == i[0], Member.isTsetPLay != 1).first()
                    if reb[0]:
                        if reb[0] == i_id:
                            n += 1
                            sunyi += i[2]
                except:
                    db.session.rollback()
                    continue
            list_b.append(i_id)
            list_b.append(sunyi)
            list_b.append(n)
            list_a.append(list_b)

        # 会员没有在KK玩过,但有存取款或者优惠
        # 查询每个会员每天的优惠  discounts
        uid_list = []
        for i in data_sunyi_copy:
            uid_list.append(i[0])

        res_discounts = db.session.query(
            MemberAccountChangeRecord.memberId

        ).filter(or_(MemberAccountChangeRecord.accountChangeType == 121,
                     MemberAccountChangeRecord.accountChangeType == 100010,
                     MemberAccountChangeRecord.accountChangeType == 100011,
                     MemberAccountChangeRecord.accountChangeType == 900006,
                     MemberAccountChangeRecord.accountChangeType == 900007),
                 MemberAccountChangeRecord.time.between(start_time, end_time)).all()

        # 存款
        res_deposits = db.session.query(Deposit.memberId).filter(
            Deposit.status == 2, Deposit.isAcdemen == 1,
            Deposit.auditTime.between(start_time, end_time)).all()

        # 　取款
        res_withdrawals = db.session.query(Withdrawal.memberId).filter(
            Withdrawal.status == 2, Withdrawal.isAcdemen == 1,
            Withdrawal.auditTime.between(start_time, end_time)).all()

        uid_set = set()
        for i in res_discounts:
            uid_set.add(i[0])
        for i in res_deposits:
            uid_set.add(i[0])
        for i in res_withdrawals:
            uid_set.add(i[0])

        for i in uid_set:
            if i not in uid_list:
                # 找出该会员对应的代理
                parent_id = db.session.query(Member.parent, Member.username).filter(Member.id == i,
                                                                                    Member.isTsetPLay != 1).first()
                reb = db.session.query(Member.commissionConfig).filter(Member.id == parent_id[0]).first()
                is_yongjin = db.session.query(ConfigYongjin.enable).filter(ConfigYongjin.id == reb[0]).first()
                if is_yongjin[0] == 1:

                    discounts = 0
                    for parent_list in list_a:
                        # reb = db.session.query(Member.commissionConfig).filter(Member.id == parent_list[0]).first()

                        res_yongjin = db.session.query(YongJinTyb).filter(YongJinTyb.Yid == reb[0]).order_by(
                            YongJinTyb.pcJine.desc()).all()
                        if reb:
                            if parent_list[0] == parent_id[0]:

                                for yj in res_yongjin:

                                    res_discounts = db.session.query(
                                        MemberAccountChangeRecord.memberId,
                                        func.sum(MemberAccountChangeRecord.amount)
                                    ).filter(or_(MemberAccountChangeRecord.accountChangeType == 121,
                                                 MemberAccountChangeRecord.accountChangeType == 100010,
                                                 MemberAccountChangeRecord.accountChangeType == 100011,
                                                 MemberAccountChangeRecord.accountChangeType == 900006,
                                                 MemberAccountChangeRecord.accountChangeType == 900007),
                                             MemberAccountChangeRecord.memberId == i,
                                             MemberAccountChangeRecord.time.between(start_time, end_time)).first()
                                    if res_discounts[0]:
                                        discounts = round(res_discounts[1] * yj.youhui / 100, 2)

                    # 存款
                    res_deposits = db.session.query(Deposit.memberId, func.sum(Deposit.applicationAmount)).filter(
                        Deposit.memberId == i, Deposit.status == 2, Deposit.isAcdemen == 1,
                        Deposit.auditTime.between(start_time, end_time)).first()
                    if res_deposits[0]:
                        deposits = res_deposits[1]
                    else:
                        deposits = 0
                    # 　取款
                    res_withdrawals = db.session.query(Withdrawal.memberId,
                                                       func.sum(Withdrawal.withdrawalAmount)).filter(
                        Withdrawal.memberId == i, Withdrawal.status == 2, Withdrawal.isAcdemen == 1,
                        Withdrawal.auditTime.between(start_time, end_time)).first()
                    if res_withdrawals[0]:
                        withdrawals = res_withdrawals[1]
                    else:
                        withdrawals = 0
                    actionTime = int(time.time())

                    member_yongjin = MemberYongjinCompute()
                    member_yongjin.recordId = recordId
                    member_yongjin.uid = i
                    member_yongjin.username = parent_id[1]
                    member_yongjin.discounts = discounts  # 优惠
                    member_yongjin.actionTime = actionTime
                    member_yongjin.startTime = startTime
                    member_yongjin.endTime = endTime
                    member_yongjin.type = 1
                    member_yongjin.ec_name = 'KK'
                    member_yongjin.childType = 1001
                    member_yongjin.deposits = deposits
                    member_yongjin.withdrawals = withdrawals
                    member_yongjin.parentId = parent_id[0]
                    member_yongjin.sunyi = 0
                    member_yongjin.amount = 0
                    member_yongjin.betAmount = 0

                    try:
                        db.session.add(member_yongjin)
                        db.session.commit()
                    except:
                        db.session.rollback()
                        db.session.remove()

        #  #根据uid去查blast_members表,查出相应佣金设定
        for i in data_sunyi_copy:
            # 找出该会员对应的代理
            parent_id = db.session.query(Member.parent).filter(Member.id == i[0], Member.isTsetPLay != 1).first()

            if parent_id:
                reb = db.session.query(Member.commissionConfig).filter(Member.id == parent_id[0]).first()
                is_yongjin = db.session.query(ConfigYongjin.enable).filter(ConfigYongjin.id == reb[0]).first()

                if is_yongjin[0] == 1:
                    for parent_list in list_a:
                        # reb = db.session.query(Member.commissionConfig).filter(Member.id == parent_list[0]).first()

                        res_yongjin = db.session.query(YongJinTyb).filter(YongJinTyb.Yid == reb[0]).order_by(
                            YongJinTyb.pcJine.desc()).all()
                        if reb:
                            if parent_list[0] == parent_id[0]:

                                amount = 0
                                discounts = 0
                                fanshui = 0
                                for yj in res_yongjin:

                                    if parent_list[1] >= yj.pcJine and parent_list[2] >= yj.yxhuiyuan:

                                        tyb = json.loads(yj.tuiyongbi)

                                        if tyb['kk']['1001'] == None:
                                            tyb = 0
                                        else:
                                            tyb = tyb['kk']['1001'] / 100
                                        amount = round(i[2] * tyb, 2)  # 佣金

                                        # 查询每个会员每天的优惠  discounts
                                        res_discounts = db.session.query(
                                            func.sum(MemberAccountChangeRecord.amount)
                                        ).filter(or_(MemberAccountChangeRecord.accountChangeType == 121,
                                                     MemberAccountChangeRecord.accountChangeType == 100010,
                                                     MemberAccountChangeRecord.accountChangeType == 100011,
                                                     MemberAccountChangeRecord.accountChangeType == 900006,
                                                     MemberAccountChangeRecord.accountChangeType == 900007),
                                                 MemberAccountChangeRecord.memberId == i[0],
                                                 MemberAccountChangeRecord.time.between(start_time, end_time)).first()
                                        if res_discounts[0]:
                                            discounts = round(res_discounts[0] * yj.youhui / 100, 2)

                                        # 批次返水
                                        res_fanshui = db.session.query(func.sum(MemberFanshuiPc.amount)).filter(
                                            MemberFanshuiPc.uid == i[0],
                                            MemberFanshuiPc.fanshuiTime.between(startTime, endTime),
                                            MemberFanshuiPc.ec_name == 'KK',
                                            MemberFanshuiPc.childType == '1001',
                                        ).first()
                                        if res_fanshui[0]:
                                            fanshui = round(res_fanshui[0] * yj.fanshui / 100, 2)

                                        break
                                    else:
                                        tyb = json.loads(res_yongjin[-1].tuiyongbi)

                                        if tyb['kk']['1001'] == None:
                                            tyb = 0
                                        else:
                                            tyb = tyb['kk']['1001'] / 100
                                        amount = round(i[2] * tyb, 2)  # 佣金

                                        # 查询每个会员每天的优惠  discounts
                                        res_discounts = db.session.query(
                                            func.sum(MemberAccountChangeRecord.amount)
                                        ).filter(or_(MemberAccountChangeRecord.accountChangeType == 121,
                                                     MemberAccountChangeRecord.accountChangeType == 100010,
                                                     MemberAccountChangeRecord.accountChangeType == 100011),
                                                 MemberAccountChangeRecord.memberId == i[0],
                                                 MemberAccountChangeRecord.time.between(start_time, end_time)).first()
                                        if res_discounts[0]:
                                            discounts = round(res_discounts[0] * yj.youhui / 100, 2)

                                        # 批次返水
                                        res_fanshui = db.session.query(func.sum(MemberFanshuiPc.amount)).filter(
                                            MemberFanshuiPc.uid == i[0],
                                            MemberFanshuiPc.fanshuiTime.between(startTime, endTime),
                                            MemberFanshuiPc.ec_name == 'KK',
                                            MemberFanshuiPc.childType == '1001'
                                        ).first()
                                        if res_fanshui[0]:
                                            fanshui = round(res_fanshui[0] * yj.fanshui / 100, 2)

                                sunyi = round(i[2], 2)  # 损益
                                # 时时返水
                                res_ss_fanshui = db.session.query(
                                    func.sum(MemberAccountChangeRecord.amount)).filter(
                                    MemberAccountChangeRecord.memberId == i[0],
                                    MemberAccountChangeRecord.accountChangeType == 2,
                                    MemberAccountChangeRecord.time.between(start_time, end_time)).first()
                                if res_ss_fanshui[0]:
                                    ss_fanshui = res_ss_fanshui[0]
                                else:
                                    ss_fanshui = 0

                                # 存款
                                res_deposits = db.session.query(func.sum(Deposit.applicationAmount)).filter(
                                    Deposit.memberId == i[0], Deposit.status == 2, Deposit.isAcdemen == 1,
                                    Deposit.auditTime.between(start_time, end_time)).first()

                                if res_deposits[0]:
                                    deposits = res_deposits[0]
                                else:
                                    deposits = 0

                                # 　取款
                                res_withdrawals = db.session.query(func.sum(Withdrawal.withdrawalAmount)).filter(
                                    Withdrawal.memberId == i[0], Withdrawal.status == 2, Withdrawal.isAcdemen == 1,
                                    Withdrawal.auditTime.between(start_time, end_time)).first()
                                if res_withdrawals[0]:
                                    withdrawals = res_withdrawals[0]
                                else:
                                    withdrawals = 0

                                # 用深拷贝
                                betAmount = 0
                                for bet in data_betAmount:
                                    if bet[0] == i[0]:
                                        betAmount = bet[2]
                                # current_app.logger.info("%s会员%s的损益%s,优惠%s,存款%s,取款%s,投注金额%s" % (
                                #     '2019-07-18', i[0], amount, discounts, deposits, withdrawals, betAmount))
                                actionTime = int(time.time())

                                member_yongjin = MemberYongjinCompute()
                                member_yongjin.recordId = recordId
                                member_yongjin.uid = i[0]
                                member_yongjin.username = i[1]
                                member_yongjin.sunyi = sunyi
                                member_yongjin.amount = amount
                                member_yongjin.discounts = discounts
                                member_yongjin.betAmount = betAmount
                                member_yongjin.actionTime = actionTime
                                member_yongjin.startTime = startTime
                                member_yongjin.endTime = endTime
                                member_yongjin.fanshui = fanshui
                                member_yongjin.ss_fanshui = ss_fanshui
                                member_yongjin.type = 1
                                member_yongjin.ec_name = 'KK'
                                member_yongjin.childType = 1001
                                member_yongjin.deposits = deposits
                                member_yongjin.withdrawals = withdrawals
                                member_yongjin.parentId = parent_id[0]

                                try:
                                    db.session.add(member_yongjin)
                                    db.session.commit()
                                except:
                                    db.session.rollback()
                                    db.session.remove()

        # 娱乐城数据加到tb_member_yongjin
        new_result_EC_betAmount = db.session.query(
            EntertainmentCityBetsDetail.ECCode,
            EntertainmentCityBetsDetail.PlayerName,
            func.sum(EntertainmentCityBetsDetail.ValidBetAmount),
            func.sum(EntertainmentCityBetsDetail.CusAccount),
            EntertainmentCityBetsDetail.childType,
        ).filter(EntertainmentCityBetsDetail.ReckonTime.between(start_time, end_time)) \
            .group_by(EntertainmentCityBetsDetail.ECCode, EntertainmentCityBetsDetail.PlayerName,
                      EntertainmentCityBetsDetail.childType).all()

        for res in new_result_EC_betAmount:

            # print(res)
            reb = db.session.query(Member.id, Member.parent).filter(Member.username == res[1]).first()
            result = db.session.query(Member.commissionConfig).filter(Member.id == reb[1],
                                                                      Member.isTsetPLay != 1).first()

            result_tyb = db.session.query(YongJinTyb).filter(YongJinTyb.Yid == result[0]).order_by(
                YongJinTyb.pcJine.desc()).all()

            is_yongjin = db.session.query(ConfigYongjin.enable).filter(ConfigYongjin.id == result[0]).first()
            if is_yongjin[0] == 1:

                amount = 0
                fanshui = 0
                for parent_list in list_a:
                    if reb[1] == parent_list[0]:
                        for res_tyb in result_tyb:
                            if parent_list[1] >= res_tyb.pcJine and parent_list[2] >= res_tyb.yxhuiyuan:
                                tyb = json.loads(res_tyb.tuiyongbi)
                                for k, v in tyb.items():
                                    if k == res[0]:
                                        for key, value in v.items():
                                            if int(key) == res[4]:
                                                if value == None:
                                                    tyb = 0
                                                else:
                                                    tyb = value
                                                amount = round((res[2] - res[3]) * tyb / 100, 2)
                                                fanshui = db.session.query(func.sum(MemberFanshuiPc.amount)).filter(
                                                    MemberFanshuiPc.username == res[1],
                                                    MemberFanshuiPc.fanshuiTime.between(startTime, endTime),
                                                    MemberFanshuiPc.ec_name == k,
                                                    MemberFanshuiPc.childType == int(key)
                                                ).first()
                                                if fanshui[0] == None:
                                                    fanshui = 0
                                                else:
                                                    fanshui = fanshui[0]

                                break
                            else:
                                tyb = json.loads(result_tyb[-1].tuiyongbi)
                                for k, v in tyb.items():
                                    if k == res[0]:
                                        for key, value in v.items():
                                            if int(key) == res[4]:
                                                if value == None:
                                                    tyb = 0
                                                else:
                                                    tyb = value
                                                amount = round((res[2] - res[3]) * tyb / 100, 2)
                                                fanshui = db.session.query(func.sum(MemberFanshuiPc.amount)).filter(
                                                    MemberFanshuiPc.username == res[1],
                                                    MemberFanshuiPc.fanshuiTime.between(startTime, endTime),
                                                    MemberFanshuiPc.ec_name == k,
                                                    MemberFanshuiPc.childType == int(key)
                                                ).first()
                                                if fanshui[0] == None:
                                                    fanshui = 0
                                                else:
                                                    fanshui = fanshui[0]

                actionTime = int(time.time())
                # yongjinTime = (datetime.date.today() + datetime.timedelta(days=-1)).strftime("%Y-%m-%d")

                ec_member_yongjin = MemberYongjinCompute()
                ec_member_yongjin.recordId = recordId
                ec_member_yongjin.uid = reb[0]
                ec_member_yongjin.username = res[1]
                ec_member_yongjin.sunyi = round(res[2] - res[3], 2)  # 损益
                ec_member_yongjin.amount = amount  # 佣金
                ec_member_yongjin.fanshui = fanshui  # 返水
                ec_member_yongjin.betAmount = res[2]
                ec_member_yongjin.actionTime = actionTime
                ec_member_yongjin.startTime = startTime
                ec_member_yongjin.endTime = endTime
                ec_member_yongjin.type = 2
                ec_member_yongjin.ec_name = res[0]
                ec_member_yongjin.childType = res[4]
                ec_member_yongjin.parentId = reb[1]
                try:
                    db.session.add(ec_member_yongjin)
                    db.session.commit()
                except:
                    db.session.rollback()
                    db.session.remove()

        # 代理佣金计算

        res_agent_yongjin = db.session.query(
            MemberYongjinCompute.parentId,
            func.sum(MemberYongjinCompute.sunyi).label("sunyi"),
            func.sum(MemberYongjinCompute.amount).label("amount"),
            func.sum(MemberYongjinCompute.discounts).label("discounts"),
            func.sum(MemberYongjinCompute.fanshui).label("fanshui"),
            func.sum(MemberYongjinCompute.deposits).label("deposits"),
            func.sum(MemberYongjinCompute.withdrawals).label("withdrawals"),
            func.sum(MemberYongjinCompute.betAmount).label("betAmount"),
            func.sum(MemberYongjinCompute.ss_fanshui).label("ss_fanshui")  # i[8]
        ).filter(MemberYongjinCompute.recordId == recordId).group_by(
            MemberYongjinCompute.parentId).all()
        # current_app.logger.info("所有代理的聚合查询的结果集%s" % res_agent_yongjin)
        # try:
        #     res_recordId = db.session.query(MemberAgentDetail.recordId).filter().order_by(
        #         MemberAgentDetail.recordId.desc()).first()
        #
        #     recordId = res_recordId[0] + 1
        # except:
        #     db.session.rollback()
        #     recordId = 1

        data_a = []
        for i in res_agent_yongjin:  # i[0]是代理id

            if i[0]:
                d = {}
                d['sunyi'] = round(i[1], 2)
                d['betAmount'] = round(i[7], 2)
                # 计算存取款手续费
                reb = db.session.query(Member.commissionConfig, Member.username, Member.type).filter(
                    Member.id == i[0]).first()

                try:
                    agent_sheding = db.session.query(ConfigYongjin).filter(ConfigYongjin.id == reb[0]).first()
                except:
                    return {'success': False, 'errorMsg': '没有有效数据数据'}
                current_app.logger.info("查询代理%s对应的佣金设定%s" % (i[0], agent_sheding))

                cun = i[5] * agent_sheding.dbcksxf / 100
                if cun < agent_sheding.cksxfsx:
                    agent_cun = cun
                else:
                    agent_cun = agent_sheding.cksxfsx

                qu = i[6] * agent_sheding.dbqksxf / 100
                if qu < agent_sheding.qksxfsx:
                    agent_qu = qu
                else:
                    agent_qu = agent_sheding.qksxfsx

                # 计算代理佣金
                agent_yj = round(i[2] - i[3] - i[4] - i[8] - (agent_cun + agent_qu), 2)
                d['agent_yj'] = agent_yj

                # 计算每个代理下的人数
                res_member_id = db.session.query(Member.id).filter(Member.parent == i[0],
                                                                   Member.isTsetPLay != 1).all()
                res_member_id_yj = db.session.query(MemberYongjinCompute.uid).filter(
                    MemberYongjinCompute.startTime == startTime, MemberYongjinCompute.endTime == endTime).all()
                member_count = set()
                for res in res_member_id:
                    for res_yj in res_member_id_yj:
                        if res[0] == res_yj[0]:
                            member_count.add(res_yj[0])

                d['member_count'] = len(member_count)
                d['username'] = reb[1]
                d['id'] = i[0]
                d['type'] = reb[2]
                d['recordId'] = recordId
                current_app.logger.info("给定时间%s内代理%s的佣金%s,会员人数%s,损益%s,有效投注金额%s" % (
                    '2019-7-30', d['id'], d['agent_yj'], d['member_count'], d['sunyi'], d['betAmount']))
                data_a.append(d)

                youhui = i[3]
                youhui_bi = 0
                fanshui = i[4]
                fanshui_bi = 0
                ss_fanshui = i[8]
                deposits = i[5]
                withdrawals = i[6]
                agent_tyb = db.session.query(YongJinTyb).filter(YongJinTyb.Yid == reb[0]).order_by(
                    YongJinTyb.pcJine.desc()).all()
                for result in agent_tyb:
                    if d['sunyi'] >= result.pcJine and d['member_count'] >= result.yxhuiyuan:
                        youhui_bi = result.youhui
                        fanshui_bi = result.fanshui
                        break
                    else:
                        youhui_bi = agent_tyb[-1].youhui
                        fanshui_bi = agent_tyb[-1].fanshui
                        break

                memberAgent = MemberAgentDetail()
                memberAgent.uid = d['id']
                memberAgent.recordId = recordId
                memberAgent.username = d['username']
                memberAgent.type = d['type']
                memberAgent.yongjin = d['agent_yj']
                memberAgent.memberCount = d['member_count']
                memberAgent.sunyi = d['sunyi']
                memberAgent.betAmount = d['betAmount']
                memberAgent.youhui = youhui
                memberAgent.youhui_bi = youhui_bi
                memberAgent.fanshui = fanshui
                memberAgent.fanshui_bi = fanshui_bi
                memberAgent.ss_fanshui = ss_fanshui
                memberAgent.deposits = deposits
                memberAgent.withdrawals = withdrawals
                memberAgent.startTime = startTime
                memberAgent.endTime = endTime

                try:
                    db.session.add(memberAgent)
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    db.session.remove()
        return make_response(data_a)


class AgentYongjinExport(Resource):
    def get(self):
        parser = RequestParser(trim=True)
        parser.add_argument('recordId', type=int)
        parser.add_argument('uid', type=int)
        args = parser.parse_args(strict=True)

        agentId = int(request.args.get('uid'))
        recordId = int(request.args.get('recordId'))

        res_agent_yj = db.session.query(MemberAgentDetail.startTime, MemberAgentDetail.endTime).filter(
            MemberAgentDetail.uid == agentId, MemberAgentDetail.recordId == recordId).first()

        res_member_yj = db.session.query(
            MemberYongjinCompute.username,
            func.sum(MemberYongjinCompute.sunyi).label("sunyi"),
            func.sum(MemberYongjinCompute.discounts).label("discounts"),
            func.sum(MemberYongjinCompute.fanshui).label("fanshui"),
            func.sum(MemberYongjinCompute.ss_fanshui).label("ss_fanshui"),
            func.sum(MemberYongjinCompute.deposits).label("deposits"),
            func.sum(MemberYongjinCompute.withdrawals).label("withdrawals"),
            func.sum(MemberYongjinCompute.betAmount).label("betAmount")).filter(
            MemberYongjinCompute.parentId == agentId,
            MemberYongjinCompute.recordId == recordId
        ).group_by(MemberYongjinCompute.username).all()

        res_member_ec = db.session.query(MemberYongjinCompute.uid, MemberYongjinCompute.username,
                                         func.sum(MemberYongjinCompute.amount).label("amount"),
                                         func.sum(MemberYongjinCompute.betAmount).label("betAmount"),
                                         MemberYongjinCompute.ec_name, MemberYongjinCompute.childType).filter(
            MemberYongjinCompute.parentId == agentId,
            MemberYongjinCompute.recordId == recordId).group_by(
            MemberYongjinCompute.id, MemberYongjinCompute.ec_name, MemberYongjinCompute.childType).all()

        from openpyxl import Workbook
        import os, time
        workbook = Workbook()
        worksheet = workbook.create_sheet("会员资料", 0)
        biaoti = ['会员名', '总损益', '总优惠', '总存款', '总提款', '总返水', '总时时返水', '总有效投注', 'KK彩票有效投注', 'KK彩票损益', 'AG视讯有效投注',
                  'AG视讯损益', 'AG电子有效投注', 'AG电子损益', 'PT电子有效投注', 'PT电子损益', 'PT捕鱼有效投注', 'PT捕鱼损益',
                  'KAIYUAN棋牌有效投注', 'KAIYUAN棋牌损益']
        worksheet.append(biaoti)

        for result in res_member_yj:
            resul_list = []
            for res in result:
                if res == None:
                    res = 0
                resul_list.append(res)

            for i in res_member_ec:
                if i.username == result.username:
                    if i.ec_name == 'KK' and i.childType == 1001:
                        if len(resul_list) == 8:
                            resul_list.append(round(i.betAmount, 2))
                            resul_list.append(round(i.amount, 2))
                        else:
                            resul_list[8] = round(i.betAmount, 2)
                            resul_list[9] = round(i.amount, 2)
                    else:
                        if len(resul_list) == 8:
                            resul_list.append(0)
                            resul_list.append(0)
                    if i.ec_name == 'AG' and i.childType == 1002:
                        if len(resul_list) == 10:
                            resul_list.append(round(i.betAmount, 2))
                            resul_list.append(round(i.amount, 2))
                        else:
                            resul_list[10] = round(i.betAmount, 2)
                            resul_list[11] = round(i.amount, 2)
                    else:
                        if len(resul_list) == 10:
                            resul_list.append(0)
                            resul_list.append(0)
                    if i.ec_name == 'AG' and i.childType == 1004:
                        if len(resul_list) == 12:
                            resul_list.append(round(i.betAmount, 2))
                            resul_list.append(round(i.amount, 2))
                        else:
                            resul_list[12] = round(i.betAmount, 2)
                            resul_list[13] = round(i.amount, 2)
                    else:
                        if len(resul_list) == 12:
                            resul_list.append(0)
                            resul_list.append(0)
                    if i.ec_name == 'PT' and i.childType == 1004:
                        if len(resul_list) == 14:
                            resul_list.append(round(i.betAmount, 2))
                            resul_list.append(round(i.amount, 2))
                        else:
                            resul_list[14] = round(i.betAmount, 2)
                            resul_list[15] = round(i.amount, 2)
                    else:
                        if len(resul_list) == 14:
                            resul_list.append(0)
                            resul_list.append(0)
                    if i.ec_name == 'PT' and i.childType == 1007:
                        if len(resul_list) == 16:
                            resul_list.append(round(i.betAmount, 2))
                            resul_list.append(round(i.amount, 2))
                        else:
                            resul_list[16] = round(i.betAmount, 2)
                            resul_list[17] = round(i.amount, 2)
                    else:
                        if len(resul_list) == 16:
                            resul_list.append(0)
                            resul_list.append(0)
                    if i.ec_name == 'KAIYUAN' and i.childType == 1003:
                        if len(resul_list) == 18:
                            resul_list.append(round(i.betAmount, 2))
                            resul_list.append(round(i.amount, 2))
                        else:
                            resul_list[18] = round(i.betAmount, 2)
                            resul_list[19] = round(i.amount, 2)
                    else:
                        if len(resul_list) == 18:
                            resul_list.append(0)
                            resul_list.append(0)

            worksheet.append(resul_list)

        filename = 'result-' + str(int(time.time())) + '.xlsx'
        workbook.save(os.path.join(current_app.static_folder, filename))
        # path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        # path = os.path.join(path, 'static')
        # workbook.save(os.path.join(path, filename))

        return make_response([{
            'success': True,
            'resultFilename': filename,
        }])


class AllAgentYongjinExport(Resource):
    def get(self):
        parser = RequestParser(trim=True)
        parser.add_argument('recordId', type=int)
        parser.add_argument('hasMemberDetail', type=int)
        args = parser.parse_args(strict=True)

        recordId = int(request.args.get('recordId'))
        hasMemberDetail = int(request.args.get('hasMemberDetail'))

        res_agent_yj = db.session.query(MemberAgentDetail.startTime, MemberAgentDetail.endTime).filter(
            MemberAgentDetail.recordId == recordId).first()

        if hasMemberDetail == 1:
            res_member_yj = db.session.query(
                MemberYongjinCompute.username,
                func.sum(MemberYongjinCompute.sunyi).label("sunyi"),
                func.sum(MemberYongjinCompute.discounts).label("discounts"),
                func.sum(MemberYongjinCompute.deposits).label("deposits"),
                func.sum(MemberYongjinCompute.withdrawals).label("withdrawals"),
                func.sum(MemberYongjinCompute.fanshui).label("fanshui"),
                func.sum(MemberYongjinCompute.ss_fanshui).label("ss_fanshui"),
                func.sum(MemberYongjinCompute.betAmount).label("betAmount")).filter(
                MemberYongjinCompute.recordId == recordId).group_by(
                MemberYongjinCompute.username).all()

            res_member_ec = db.session.query(MemberYongjinCompute.uid, MemberYongjinCompute.username,
                                             func.sum(MemberYongjinCompute.amount).label("amount"),
                                             func.sum(MemberYongjinCompute.betAmount).label("betAmount"),
                                             MemberYongjinCompute.ec_name, MemberYongjinCompute.childType).filter(
                MemberYongjinCompute.recordId == recordId).group_by(
                MemberYongjinCompute.id, MemberYongjinCompute.ec_name, MemberYongjinCompute.childType).all()

            from openpyxl import Workbook
            import os, time
            workbook = Workbook()
            ws1 = workbook.create_sheet("会员资料", 0)
            biaoti = ['会员名', '总损益', '总优惠', '总存款', '总提款', '总返水', '总时时返水', '总有效投注', 'KK彩票有效投注', 'KK彩票损益', 'AG视讯有效投注',
                      'AG视讯损益', 'AG电子有效投注', 'AG电子损益', 'PT电子有效投注', 'PT电子损益', 'PT捕鱼有效投注', 'PT捕鱼损益',
                      'KAIYUAN棋牌有效投注', 'KAIYUAN棋牌损益']
            ws1.append(biaoti)

            for result in res_member_yj:
                resul_list = []
                for res in result:
                    if res == None:
                        res = 0
                    resul_list.append(res)

                for i in res_member_ec:
                    if i.username == result.username:
                        if i.ec_name == 'KK' and i.childType == 1001:
                            if len(resul_list) == 8:
                                resul_list.append(round(i.betAmount, 2))
                                resul_list.append(round(i.amount, 2))
                            else:
                                resul_list[8] = round(i.betAmount, 2)
                                resul_list[9] = round(i.amount, 2)
                        else:
                            if len(resul_list) == 8:
                                resul_list.append(0)
                                resul_list.append(0)
                        if i.ec_name == 'AG' and i.childType == 1002:
                            if len(resul_list) == 10:
                                resul_list.append(round(i.betAmount, 2))
                                resul_list.append(round(i.amount, 2))
                            else:
                                resul_list[10] = round(i.betAmount, 2)
                                resul_list[11] = round(i.amount, 2)
                        else:
                            if len(resul_list) == 10:
                                resul_list.append(0)
                                resul_list.append(0)
                        if i.ec_name == 'AG' and i.childType == 1004:
                            if len(resul_list) == 12:
                                resul_list.append(round(i.betAmount, 2))
                                resul_list.append(round(i.amount, 2))
                            else:
                                resul_list[12] = round(i.betAmount, 2)
                                resul_list[13] = round(i.amount, 2)
                        else:
                            if len(resul_list) == 12:
                                resul_list.append(0)
                                resul_list.append(0)
                        if i.ec_name == 'PT' and i.childType == 1004:
                            if len(resul_list) == 14:
                                resul_list.append(round(i.betAmount, 2))
                                resul_list.append(round(i.amount, 2))
                            else:
                                resul_list[14] = round(i.betAmount, 2)
                                resul_list[15] = round(i.amount, 2)
                        else:
                            if len(resul_list) == 14:
                                resul_list.append(0)
                                resul_list.append(0)
                        if i.ec_name == 'PT' and i.childType == 1007:
                            if len(resul_list) == 16:
                                resul_list.append(round(i.betAmount, 2))
                                resul_list.append(round(i.amount, 2))
                            else:
                                resul_list[16] = round(i.betAmount, 2)
                                resul_list[17] = round(i.amount, 2)
                        else:
                            if len(resul_list) == 16:
                                resul_list.append(0)
                                resul_list.append(0)
                        if i.ec_name == 'KAIYUAN' and i.childType == 1003:
                            if len(resul_list) == 18:
                                resul_list.append(round(i.betAmount, 2))
                                resul_list.append(round(i.amount, 2))
                            else:
                                resul_list[18] = round(i.betAmount, 2)
                                resul_list[19] = round(i.amount, 2)
                        else:
                            if len(resul_list) == 18:
                                resul_list.append(0)
                                resul_list.append(0)

                ws1.append(resul_list)

            # filename = 'result-' + str(int(time.time())) + '.xlsx'
            # # workbook.save(os.path.join(current_app.static_folder, filename))
            # path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            # path = os.path.join(path, 'static')
            # workbook.save(os.path.join(path, filename))

            res_pt_yingjin = db.session.query(
                MemberYongjinCompute.parentId,
                func.sum(MemberYongjinCompute.amount).label("amount")).filter(
                MemberYongjinCompute.recordId == recordId).group_by(
                MemberYongjinCompute.parentId).all()

            # 代理
            result_agent_yj = db.session.query(MemberAgentDetail.uid,
                                               MemberAgentDetail.username,
                                               MemberAgentDetail.type,
                                               MemberAgentDetail.yongjin,
                                               MemberAgentDetail.memberCount,
                                               MemberAgentDetail.sunyi,
                                               MemberAgentDetail.youhui,
                                               MemberAgentDetail.fanshui,
                                               MemberAgentDetail.ss_fanshui,
                                               MemberAgentDetail.deposits,
                                               MemberAgentDetail.withdrawals,
                                               MemberAgentDetail.betAmount
                                               ).filter(MemberAgentDetail.recordId == recordId).all()

            from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
            ws2 = workbook.create_sheet("代理佣金", 1)

            ws2.merge_cells('A1:A2')
            ws2.merge_cells('B1:B2')
            ws2.merge_cells('C1:C2')
            ws2.merge_cells('D1:D2')
            ws2.merge_cells('E1:E2')
            ws2.merge_cells('F1:F2')
            ws2.merge_cells('G1:I1')
            ws2.merge_cells('J1:L1')
            ws2.merge_cells('M1:O1')
            ws2.merge_cells('P1:P2')
            ws2.merge_cells('Q1:Q2')
            ws2.merge_cells('R1:R2')
            ws2.merge_cells('S1:S2')
            ws2.merge_cells('T1:T2')
            ws2['A1'] = "层级"
            ws2['B1'] = "账号"
            ws2['C1'] = "佣金"
            ws2['D1'] = "有效会员"
            ws2['E1'] = "总损益"
            ws2['F1'] = "平台佣金"
            ws2['G1'] = "优惠"
            ws2['J1'] = "返水"
            ws2['M1'] = "时时返水"
            ws2['G2'] = "总优惠"
            ws2['H2'] = "优惠占比"
            ws2['I2'] = "优惠负担额"
            ws2['J2'] = "总返水"
            ws2['K2'] = "返水占比"
            ws2['L2'] = "返水负担额"
            ws2['M2'] = "总时时返水"
            ws2['N2'] = "时时返水占比"
            ws2['O2'] = "时时返水负担额"
            ws2['P1'] = "总存款"
            ws2['Q1'] = "总存款手续费"
            ws2['R1'] = "总取款"
            ws2['S1'] = "总取款手续费"
            ws2['T1'] = "总有效投注"
            ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws2["B1"].alignment = Alignment(horizontal="center", vertical="center")
            ws2["C1"].alignment = Alignment(horizontal="center", vertical="center")
            ws2["D1"].alignment = Alignment(horizontal="center", vertical="center")
            ws2["E1"].alignment = Alignment(horizontal="center", vertical="center")
            ws2["F1"].alignment = Alignment(horizontal="center", vertical="center")
            ws2["G1"].alignment = Alignment(horizontal="center", vertical="center")
            ws2["J1"].alignment = Alignment(horizontal="center", vertical="center")
            ws2["M1"].alignment = Alignment(horizontal="center", vertical="center")
            ws2["P1"].alignment = Alignment(horizontal="center", vertical="center")
            ws2["Q1"].alignment = Alignment(horizontal="center", vertical="center")
            ws2["R1"].alignment = Alignment(horizontal="center", vertical="center")
            ws2["S1"].alignment = Alignment(horizontal="center", vertical="center")
            ws2["T1"].alignment = Alignment(horizontal="center", vertical="center")

            for result in result_agent_yj:
                result_bi = db.session.query(MemberAgentDetail.youhui_bi, MemberAgentDetail.fanshui_bi).filter(
                    MemberAgentDetail.recordId == recordId, MemberAgentDetail.uid == result.uid).first()

                youhui_bi = '%d%%' % result_bi[0]
                fanshui_bi = '%d%%' % result_bi[1]
                ss_fanshui_bi = '%d%%' % 100

                res_list = []
                reb = db.session.query(Member.commissionConfig).filter(Member.id == result[0]).first()
                # r = db.session.query(YongJinTyb).filter(YongJinTyb.Yid==reb[0]).order_by(YongJinTyb.pcJine.desc()).all()
                agent_sheding = db.session.query(ConfigYongjin).filter(ConfigYongjin.id == reb[0]).first()
                cksxf = result[9] * agent_sheding.dbcksxf / 100
                qksxf = result[10] * agent_sheding.dbqksxf / 100
                pt_yongjin = None  # 平台佣金
                for re in res_pt_yingjin:
                    r_name = db.session.query(Member.username).filter(Member.id == re[0]).first()
                    if result.username == r_name[0]:
                        pt_yongjin = round(re[1], 2)

                res_list.append(result[2])
                res_list.append(result[1])
                res_list.append(result[3])
                res_list.append(result[4])
                res_list.append(result[5])  # 总损益
                res_list.append(pt_yongjin)
                res_list.append(result[6])
                res_list.append(youhui_bi)
                res_list.append(result[6])
                res_list.append(result[7])
                res_list.append(fanshui_bi)
                res_list.append(result[7])
                res_list.append(result[8])
                res_list.append(ss_fanshui_bi)
                res_list.append(result[8])
                res_list.append(result[9])
                res_list.append(cksxf)
                res_list.append(result[10])
                res_list.append(qksxf)
                res_list.append(result[11])
                ws2.append(res_list)
                # print(res_list)

            filename = 'result-' + str(int(time.time())) + '.xlsx'
            workbook.save(os.path.join(current_app.static_folder, filename))
            # path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            # path = os.path.join(path, 'static')
            # workbook.save(os.path.join(path, filename))
            return make_response([{
                'success': True,
                'resultFilename': filename,
            }])

        if hasMemberDetail == 0:
            res_pt_yingjin = db.session.query(
                MemberYongjinCompute.parentId,
                func.sum(MemberYongjinCompute.amount).label("amount")).filter(
                MemberYongjinCompute.recordId == recordId).group_by(
                MemberYongjinCompute.parentId).all()
            # 代理
            result_agent_yj = db.session.query(MemberAgentDetail.uid,
                                               MemberAgentDetail.username,
                                               MemberAgentDetail.type,
                                               MemberAgentDetail.yongjin,
                                               MemberAgentDetail.memberCount,
                                               MemberAgentDetail.sunyi,
                                               MemberAgentDetail.youhui,
                                               MemberAgentDetail.fanshui,
                                               MemberAgentDetail.ss_fanshui,
                                               MemberAgentDetail.deposits,
                                               MemberAgentDetail.withdrawals,
                                               MemberAgentDetail.betAmount
                                               ).filter(MemberAgentDetail.recordId == recordId).all()

            from openpyxl import Workbook
            import os, time
            from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
            workbook_agent = Workbook()
            ws3 = workbook_agent.create_sheet("代理佣金", 0)
            ws3.merge_cells('A1:A2')
            ws3.merge_cells('B1:B2')
            ws3.merge_cells('C1:C2')
            ws3.merge_cells('D1:D2')
            ws3.merge_cells('E1:E2')
            ws3.merge_cells('F1:F2')
            ws3.merge_cells('G1:I1')
            ws3.merge_cells('J1:L1')
            ws3.merge_cells('M1:O1')
            ws3.merge_cells('P1:P2')
            ws3.merge_cells('Q1:Q2')
            ws3.merge_cells('R1:R2')
            ws3.merge_cells('S1:S2')
            ws3.merge_cells('T1:T2')
            ws3['A1'] = "层级"
            ws3['B1'] = "账号"
            ws3['C1'] = "佣金"
            ws3['D1'] = "有效会员"
            ws3['E1'] = "总损益"
            ws3['F1'] = "平台佣金"
            ws3['G1'] = "优惠"
            ws3['J1'] = "返水"
            ws3['M1'] = "时时返水"
            ws3['G2'] = "总优惠"
            ws3['H2'] = "优惠占比"
            ws3['I2'] = "优惠负担额"
            ws3['J2'] = "总返水"
            ws3['K2'] = "返水占比"
            ws3['L2'] = "返水负担额"
            ws3['M2'] = "总时时返水"
            ws3['N2'] = "时时返水占比"
            ws3['O2'] = "时时返水负担额"
            ws3['P1'] = "总存款"
            ws3['Q1'] = "总存款手续费"
            ws3['R1'] = "总取款"
            ws3['S1'] = "总取款手续费"
            ws3['T1'] = "总有效投注"
            ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws3["B1"].alignment = Alignment(horizontal="center", vertical="center")
            ws3["C1"].alignment = Alignment(horizontal="center", vertical="center")
            ws3["D1"].alignment = Alignment(horizontal="center", vertical="center")
            ws3["E1"].alignment = Alignment(horizontal="center", vertical="center")
            ws3["F1"].alignment = Alignment(horizontal="center", vertical="center")
            ws3["G1"].alignment = Alignment(horizontal="center", vertical="center")
            ws3["J1"].alignment = Alignment(horizontal="center", vertical="center")
            ws3["M1"].alignment = Alignment(horizontal="center", vertical="center")
            ws3["P1"].alignment = Alignment(horizontal="center", vertical="center")
            ws3["Q1"].alignment = Alignment(horizontal="center", vertical="center")
            ws3["R1"].alignment = Alignment(horizontal="center", vertical="center")
            ws3["S1"].alignment = Alignment(horizontal="center", vertical="center")
            ws3["T1"].alignment = Alignment(horizontal="center", vertical="center")

            # biaoti_agent = ['会员等级', '账号', '佣金', '有效会员', '总损益', '平台佣金', '总优惠', '优惠占比', '优惠负担额', '总返水', '总返水占比', '总返水负担额',
            #                 '总时时返水', '总时时返水占比', '总时时返水负担额', '总存款', '总存款手续费', '总提款', '总提款手续费', '总有效投注', ]
            # worksheet_agent.append(biaoti_agent)

            for result in result_agent_yj:
                result_bi = db.session.query(MemberAgentDetail.youhui_bi, MemberAgentDetail.fanshui_bi).filter(
                    MemberAgentDetail.recordId == recordId, MemberAgentDetail.uid == result.uid).first()

                youhui_bi = '%d%%' % result_bi[0]
                fanshui_bi = '%d%%' % result_bi[1]
                ss_fanshui_bi = '%d%%' % 100

                res_list = []
                reb = db.session.query(Member.commissionConfig).filter(Member.id == result[0]).first()
                # r = db.session.query(YongJinTyb).filter(YongJinTyb.Yid==reb[0]).order_by(YongJinTyb.pcJine.desc()).all()
                agent_sheding = db.session.query(ConfigYongjin).filter(ConfigYongjin.id == reb[0]).first()
                cksxf = result[9] * agent_sheding.dbcksxf / 100
                qksxf = result[10] * agent_sheding.dbqksxf / 100
                pt_yongjin = None  # 平台佣金
                for re in res_pt_yingjin:
                    r_name = db.session.query(Member.username).filter(Member.id == re[0]).first()
                    if result.username == r_name[0]:
                        pt_yongjin = round(re[1], 2)

                res_list.append(result[2])
                res_list.append(result[1])
                res_list.append(result[3])
                res_list.append(result[4])
                res_list.append(result[5])  # 总损益
                res_list.append(pt_yongjin)
                res_list.append(result[6])
                res_list.append(youhui_bi)
                res_list.append(result[6])
                res_list.append(result[7])
                res_list.append(fanshui_bi)
                res_list.append(result[7])
                res_list.append(result[8])
                res_list.append(ss_fanshui_bi)
                res_list.append(result[8])
                res_list.append(result[9])
                res_list.append(cksxf)
                res_list.append(result[10])
                res_list.append(qksxf)
                res_list.append(result[11])
                ws3.append(res_list)

            filename = 'result-' + str(int(time.time())) + '.xlsx'
            workbook_agent.save(os.path.join(current_app.static_folder, filename))
            # path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            # path = os.path.join(path, 'static')
            # workbook_agent.save(os.path.join(path, filename))

            return make_response([{
                'success': True,
                'resultFilename': filename,
            }])
