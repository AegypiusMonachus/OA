import time, datetime, math
import requests,base64,pyDes,binascii
from flask import current_app
def time_to_value():
	return int(time.time())


def value_to_time(value):
	return time.ctime(value)


def host_to_value(host):
	arr = host.split('.')
	result = 0
	result += int(arr[0]) << 24
	result += int(arr[1]) << 16
	result += int(arr[2]) << 8
	result += int(arr[3])
	return result


def value_to_host(value):
	result = []
	result.append(str((value & 0xff000000) >> 24))
	result.append(str((value & 0x00ff0000) >> 16))
	result.append(str((value & 0x0000ff00) >> 8))
	result.append(str((value & 0x000000ff)))
	return '.'.join(result)


def today_lower_limit():
	today = datetime.datetime.now()
	lower = datetime.datetime.combine(today, datetime.time.min)
	lower = lower.timetuple()
	return int(time.mktime(lower))


def today_upper_limit():
	today = datetime.datetime.now()
	upper = datetime.datetime.combine(today, datetime.time.max)
	upper = upper.timetuple()
	return int(time.mktime(upper))


def analysisXLS(m_fileURL):
	from openpyxl import load_workbook
	workbook_ = load_workbook(m_fileURL)
	sheet = workbook_.active
	len_row = len(list(sheet.rows))
	m_list = []
	for row in range(2, len_row+1):
		sheet.cell(row=row, column=6).value = '存入成功'
		m_dict = {}
		m_dict['username'] = sheet.cell(row=row, column=1).value
		m_dict['coin'] = sheet.cell(row=row, column=2).value
		m_dict['fcoin'] = sheet.cell(row=row, column=3).value
		m_list.append(m_dict)
	workbook_.save(m_fileURL)
	return m_list


#计算两天时间差
def dayLimit(m_stime, m_etime):
	m_days = []
	m_days.append(m_etime)
	#stime = datetime.datetime.utcfromtimestamp(m_stime)
	#etime = datetime.datetime.utcfromtimestamp(m_etime)
	m_limit = int((m_etime-m_stime)/3600/24)
	for i in range(m_limit):
		m_days.append(m_etime - (3600*24*(i+1)))
	return m_days


def kaijiang(args):
	m_node_url = 'http://www2.devqp.info:8800/data/kj'
	m_ekey = "www.skd.com"
	m_time = args['time']
	timeArray = time.localtime(m_time)
	otherStyleTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
	m_para = 'type=%s&number=%s&time=%s&data=%s&key=%s'%(args['type'],args['actionNo'],otherStyleTime,args['data'],m_ekey)
	args['key'] = m_ekey
	url = m_node_url
	response = requests.post(url, data = m_para)
	return response.text    


def str_to_hex(mstr):
	return ''.join([hex(ord(c)).replace('0x', '') for c in mstr])


def hex_to_lid(mhex):
	m_hex = ''
	for i in range(0,len(mhex),2):
		m_str = (mhex[i:i+2])
		m_hex += chr(int(m_str, 16))
	return m_hex


def myxor(mstr,mkey):
	if (mstr is None) or (mstr == ''):
		return ''
	if mkey is None or mkey == '':
		mkey = 'lm'
	mlen1 = len(mstr)
	mlen2 = len(mkey)
	mlimit = math.ceil(mlen1/mlen2)
	for limit in range(0,mlimit-1):
		mkey += mkey
	return ''.join([chr(ord(c1) ^ ord(c2)) for c1, c2 in zip(mstr, mkey)])

def strToBase64str(data):
	b64 = base64.b64encode(data.encode('utf-8'))
	return str(b64,'utf-8')


def DesEncrypt(data,key):
	k=pyDes.des(key,pyDes.CBC,key,pad=None,padmode=pyDes.PAD_PKCS5)
	d=k.encrypt(data)
	d=str(binascii.b2a_hex(d),'utf-8')
	return d
def DesDecrypt(data,key):
	k=pyDes.des(key,pyDes.CBC,key,pad=None,padmode=pyDes.PAD_PKCS5)
	data=base64.decodestring(data)
	d=k.decrypt(data)
	return d

def dateToAmericaTime(dateStr):
	return (datetime.datetime.strptime(dateStr, "%Y-%m-%d %H:%M:%S")+datetime.timedelta(hours=-12)).strftime("%Y-%m-%d %H:%M:%S")
	
def AmericaTimeToDateTime(dateStr):
	date1 = (datetime.datetime.strptime(dateStr, "%Y-%m-%dT%H:%M:%S")+datetime.timedelta(hours=12)).strftime("%Y-%m-%dT%H:%M:%S")
	return time.mktime(time.strptime(date1, "%Y-%m-%dT%H:%M:%S"))

def defaultDateTime(dateStr):
	date1 = (datetime.datetime.strptime(dateStr, "%Y-%m-%dT%H:%M:%S")).strftime("%Y-%m-%dT%H:%M:%S")
	return time.mktime(time.strptime(date1, "%Y-%m-%dT%H:%M:%S"))

def doPost(url,data,headers):
	response  = requests.post(url,data = data,headers = headers)
	return response

def doGet(url,data,headers):
	response  = requests.get(url,params = data,headers = headers)
	return response

