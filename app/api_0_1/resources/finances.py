from flask import request, g, current_app
from flask_restful import Resource, marshal_with, fields, abort
from flask_restful.reqparse import RequestParser
from app.common.orderUtils import createOrderIdNew
from app.models import db
from app.models.common.utils import *
from app.models.finance import (
	RebateLog,
	RebateLogDetail,
	RebateDetail,
	CommissionLog,
	ImportDiscountLog
)
from app.models.member import Member
from app.common.utils import *
from ..common import *
from ..common.utils import *
from app.models.member_account_change import Withdrawal, Deposit, MemberAccountChangeRecord

class RebateLogs(Resource):

	@marshal_with(make_marshal_fields({
		'id': fields.Integer,
		'name': fields.String,
		'time': fields.Integer,
		'rebateTimeLower': fields.Integer,
		'rebateTimeUpper': fields.Integer,
		'agentUsername': fields.String,
		'operatorUsername': fields.String,
		'resultFilename': fields.String,
		'totalMember': fields.Integer,
		'totalAmount': fields.Float,
	}))
	def get(self):
		parser = RequestParser(trim=True)
		parser.add_argument('page', type=int, default=DEFAULT_PAGE)
		parser.add_argument('pageSize', type=int, default=DEFAULT_PAGE_SIZE)
		args = parser.parse_args(strict=True)

		pagination = paginate(RebateLog.query, page=args['page'], per_page=args['pageSize'])
		for item in pagination.items:
			item.totalMember = 0
			item.totalAmount = 0
			details = item.details.all()
			for detail in details:
				if detail.rebateStatus:
					item.totalMember += 1
					item.totalAmount += detail.rebateAmount
		return make_response_from_pagination(pagination)

class RebateLogDetails(Resource):

	@marshal_with(make_marshal_fields({
		'id': fields.Integer,
		'name': fields.String,
		'time': fields.Integer,
		'rebateTimeLower': fields.Integer,
		'rebateTimeUpper': fields.Integer,
		'agentUsername': fields.String,
		'operatorUsername': fields.String,
		'resultFilename': fields.String,
		'members': fields.List(fields.Nested({
			'memberId': fields.Integer,
			'memberUsername': fields.String,
			'rebateAmount': fields.Float,
			'rebateStatus': fields.Integer,
		})),
		'totalMember': fields.Integer,
		'totalAmount': fields.Float,
	}))
	def get(self, id):
		log = RebateLog.query.get_or_404(id)
		log.totalMember = 0
		log.totalAmount = 0
		log.members = []
		details = log.details.all()
		for detail in details:
			log.members.append({
				'memberId': detail.memberId,
				'memberUsername': Member.query.get(detail.memberId).username,
				'rebateAmount': detail.rebateAmount,
				'rebateStatus': detail.rebateStatus,
			})
			if detail.rebateStatus:
				log.totalMember += 1
				log.totalAmount += detail.rebateAmount
		return make_response([log])

	def post(self):
		parser = RequestParser(trim=True)
		parser.add_argument('timeLower', type=int, required=True)
		parser.add_argument('timeUpper', type=int, required=True)
		parser.add_argument('agentUsername', type=str)
		args = parser.parse_args(strict=True)

		try:
			log = RebateLog()
			log.name = ''
			log.name += value_to_time(args['timeLower'])
			log.name += ' '
			log.name += value_to_time(args['timeUpper'])
			log.time = time_to_value()
			log.rebateTimeLower = args['timeLower']
			log.rebateTimeUpper = args['timeUpper']
			log.operatorUsername = g.current_user.username
			log.resultFilename = 'result.xlsx'

			db.session.add(log)
			db.session.commit()
			try:
				members = Member.query.filter(Member.type == 0).all()
				for member in members:
					detail = RebateLogDetail()
					detail.rebateId = log.id
					detail.memberId = member.id
					detail.rebateAmount = 0
					detail.rebateStatus = 0
					log.details.append(detail)

				db.session.add(log)
				db.session.commit()
			except:
				db.session.rollback()
				db.session.remove()
				abort(500)
		except:
			db.session.rollback()
			db.session.remove()
			abort(500)

		return make_response([log.id]), 201

	def put(self, id):
		parser = RequestParser(trim=True)
		parser.add_argument('memberId', type=str, required=True)
		args = parser.parse_args(strict=True)
		return make_response([])

class RebateDetails(Resource):

	@marshal_with(make_marshal_fields({
		'memberUsername': fields.String,
		'supplierName': fields.String,
		'gameName': fields.String,
		'betAmount': fields.Float,
		'rebateAmount': fields.Float
	}, totalRebateAmount=fields.Float))
	def get(self):
		parser = RequestParser(trim=True)
		parser.add_argument('memberUsername', type=str)
		parser.add_argument('timeLower', type=int)
		parser.add_argument('timeUpper', type=int)
		args = parser.parse_args(strict=True)

		criterion = set()
		if args['memberUsername']:
			criterion.add(RebateDetail.memberUsername == args['memberUsername'])
		if args['timeLower']:
			criterion.add(RebateDetail.time > args['timeLower'])
		if args['timeUpper']:
			criterion.add(RebateDetail.time < args['timeUpper'] + SECONDS_PER_DAY)

		details = RebateDetail.query.filter(*criterion).all()
		total_rebate_amount = 0
		for detail in details:
			total_rebate_amount += detail.rebateAmount
		return make_response(details, totalRebateAmount=total_rebate_amount)

class CommissionLogs(Resource):

	@marshal_with(make_marshal_fields({
		'id': fields.Integer,
		'time': fields.Integer,
		'commissionTimeLower': fields.Integer,
		'commissionTimeUpper': fields.Integer,
		'agentUsername': fields.String,
		'numberOfChildren': fields.Integer,
		'numberOfValidBet': fields.Float,
		'commission': fields.Float,
		'profitAndLoss': fields.Float,
		'operatorUsername': fields.String,
		'resultFilename': fields.String,
	}))
	def get(self):
		parser = RequestParser(trim=True)
		parser.add_argument('page', type=int, default=DEFAULT_PAGE)
		parser.add_argument('pageSize', type=int, default=DEFAULT_PAGE_SIZE)
		parser.add_argument('agentUsername', type=str)
		parser.add_argument('timeLower', type=int)
		parser.add_argument('timeUpper', type=int)
		args = parser.parse_args(strict=True)

		pagination = paginate(CommissionLog.query, page=args['page'], per_page=args['pageSize'])
		return make_response_from_pagination(pagination)

	def post(self):
		parser = RequestParser(trim=True)
		parser.add_argument('timeLower', type=int, required=True)
		parser.add_argument('timeUpper', type=int, required=True)
		args = parser.parse_args(strict=True)

		try:
			agents = Member.query.filter(Member.type == 1).all()
			for agent in agents:
				log = CommissionLog()
				log.time = time_to_value()
				log.commissionTimeLower = args['timeLower']
				log.commissionTimeUpper = args['timeUpper']
				log.agentUsername = agent.username
				log.resultFilename = 'result.xlsx'
				log.operatorUsername = g.current_user.username

				db.session.add(log)
			db.session.commit()
		except:
			db.session.rollback()
			db.session.remove()
			abort(500)
		return make_response([]), 201

class ExportDepositAndWithdrawal(Resource):
	def get(self):
		parser = RequestParser(trim=True)
		parser.add_argument('timeLower', type=int)
		parser.add_argument('timeUpper', type=int)
		args = parser.parse_args(strict=True)
		args['timeUpper'] = args['timeUpper'] + SECONDS_PER_DAY
		m_sql = '''
				select bm.username, 
				(select levelName from blast_member_level where bm.grade = id),
				(select count(id) from blast_member_recharge bmr where bmr.uid = bm.uid and type = 100003 and state = 2 and bmr.auditTime >={0} and bmr.auditTime <{1} and isAcdemen = 1) gsrkCount,
				(select sum(amount) from blast_member_recharge bmr where bmr.uid = bm.uid and type = 100003 and state = 2 and bmr.auditTime >={0} and bmr.auditTime <{1} and isAcdemen = 1) gsrkAmount,
				(select Max(amount) from blast_member_recharge bmr where bmr.uid = bm.uid and type = 100003 and state = 2 and bmr.auditTime >={0} and bmr.auditTime <{1} and isAcdemen = 1) gsrkMaxAmount,				
				(select count(id) from blast_member_recharge bmr where bmr.uid = bm.uid and type = 100004 and state = 2 and bmr.auditTime >={0} and bmr.auditTime <{1} and isAcdemen = 1) xszfCount,
				(select sum(amount) from blast_member_recharge bmr where bmr.uid = bm.uid and type = 100004 and state = 2 and bmr.auditTime >={0} and bmr.auditTime <{1} and isAcdemen = 1) xszfAmount,
				(select Max(amount) from blast_member_recharge bmr where bmr.uid = bm.uid and type = 100004 and state = 2 and bmr.auditTime >={0} and bmr.auditTime <{1} and isAcdemen = 1) xszfMaxAmount,
				(select count(id) from blast_member_recharge bmr where bmr.uid = bm.uid and type = 900001 and state = 2 and bmr.auditTime >={0} and bmr.auditTime <{1} and isAcdemen = 1) rgrkCount,
				(select sum(amount) from blast_member_recharge bmr where bmr.uid = bm.uid and type = 900001 and state = 2 and bmr.auditTime >={0} and bmr.auditTime <{1} and isAcdemen = 1) rgrkAmount,
				(select Max(amount) from blast_member_recharge bmr where bmr.uid = bm.uid and type = 900001 and state = 2 and bmr.auditTime >={0} and bmr.auditTime <{1} and isAcdemen = 1) rgrkMaxAmount,
				(select sum(coin) from blast_coin_log bcl where bcl.uid = bm.uid and liqType in (100010,100011) and bcl.actionTime >={0} and bcl.actionTime <{1}) yh,
				(select sum(coin) from blast_coin_log bcl where bcl.uid = bm.uid and liqType = 2 and bcl.actionTime >={0} and bcl.actionTime <{1}) pcfs,
				(select sum(coin) from blast_coin_log bcl where bcl.uid = bm.uid and liqType = 122 and bcl.actionTime >={0} and bcl.actionTime <{1}) ssfs,
				(select sum(sxFee+yhFee+xzFee) from blast_member_cash bmc where bmc.uid = bm.uid and type in (200001,900002) and state = 2 and bmc.auditTime >={0} and bmc.auditTime <{1} and isAcdemen = 1) sxf, 
				(select count(id) from blast_member_cash bmc where bmc.uid = bm.uid and type in (200001,900002) and state = 2 and bmc.auditTime >={0} and bmc.auditTime <{1} and isAcdemen = 1) qkCount,
				(select sum(amount) from blast_member_cash bmc where bmc.uid = bm.uid and type in (200001,900002) and state = 2 and bmc.auditTime >={0} and bmc.auditTime <{1} and isAcdemen = 1) qkAmount	 
				from blast_members bm where bm.isTsetPLay = 0;
				'''.format(args['timeLower'],args['timeUpper'])

		from openpyxl import Workbook
		import os, time
		workbook = Workbook()
		worksheet = workbook.active
		biaoti = ['会员账号', '会员等级', '公司入款存款次数', '公司入款总额', '公司入款区间最高金额', '线上支付存款次数', '线上支付总额',
				  '线上支付区间最高金额', '人工存入次数', '人工存入总额', '人工存入区间最高金额', '优惠总额', '批次返水总额', '时时返水总额',
				  '手续费总额', '取款次数', '取款总额']
		worksheet.append(biaoti)

		m_res = db.session.execute(m_sql)
		for result in m_res:
			resul_list = []
			for res in result:
				if res == None:
					res = 0
				resul_list.append(res)
			worksheet.append(resul_list)

		filename = '总存取款-' + str(int(time.time())) + '.xlsx'
		workbook.save(os.path.join(current_app.static_folder, filename))
		return make_response([{
			'success': True,
			'resultFilename': filename,
		}])

class ImportDiscountLogs(Resource):

	@marshal_with(make_marshal_fields({
		'id': fields.Integer,
		'originalFilename': fields.String,
		'status': fields.Integer,
		'timeLower': fields.Integer,
		'timeUpper': fields.Integer,
		'operatorUsername': fields.String,
		'resultFilename': fields.String,
		'failcount': fields.Integer,
	}))
	def get(self):
		parser = RequestParser(trim=True)
		parser.add_argument('page', type=int, default=DEFAULT_PAGE)
		parser.add_argument('pageSize', type=int, default=DEFAULT_PAGE_SIZE)
		args = parser.parse_args(strict=True)

		pagination = paginate(ImportDiscountLog.query.order_by(ImportDiscountLog.id.desc()),
							  page=args['page'], per_page=args['pageSize'])
		return make_response_from_pagination(pagination)

	def post(self):
		from openpyxl import load_workbook
		import os
		parser = RequestParser(trim=True)
		parser.add_argument('fileName', type=str, required=True, location=['form','json','args'])
		args = parser.parse_args(strict=True)
		path = os.path.join(current_app.static_folder, args['fileName'])
		workbook = load_workbook(path)
		worksheet = workbook.active
		title = worksheet.delete_rows(0)
		worksheet.insert_cols(0)
		begin_time = time_to_value()
		count = 0

		for row in worksheet.rows:
			try:
				member = Member.query.filter(Member.username == row[1].value).first()
				uid = member.id
				orderId = createOrderIdNew(uid=uid)
				if not member:
					row[0].value = '未汇入'
					count += 1
					continue


				if not is_float(row[2].value):
					row[0].value = '未汇入'
					count += 1
					continue
				member_coin_log = {}
				member_coin_log['memberId'] = member.id
				member_coin_log['memberBalance'] = member.balance
				member_coin_log['memberFrozenBalance'] = member.frozenBalance
				member_coin_log['amount'] = row[2].value
				member_coin_log['accountChangeType'] = 900006
				member_coin_log['time'] = time_to_value()
				member_coin_log['host'] = host_to_value(request.remote_addr)
				member_coin_log['auditType'] = 3
				member_coin_log['isAcdemen'] = 0
				member_coin_log['orderId'] = orderId
				member_coin_log['rechargeid'] = orderId
				member_coin_log['actionUID'] = g.current_user.id

				if not is_float(row[3].value):
					row[0].value = '未汇入'
					count += 1
					continue
				member_coin_log['auditCharge'] = row[3].value
				# 备注尚未处理
				# 前台备注尚未处理

				try:
					# member_recharge_args = Deposit(**member_recharge_args)
					# db.session.add(member_recharge_args)
					member.balance += float(row[2].value)
					member.hitCodeNeed += float(row[3].value)
					member_coin_log = MemberAccountChangeRecord(**member_coin_log)
					db.session.add(member_coin_log)
					db.session.commit()
					row[0].value = '已汇入'
				except:
					db.session.rollback()
					db.session.remove()
					abort(500)
					row[0].value = '未汇入'
			except:
				return make_response(error_code=400)

		filename = 'ImportDiscountTemplate-' + str(int(time.time())) + '.xlsx'
		worksheet.insert_rows(title)
		workbook.save(os.path.join(current_app.static_folder, filename))
		end_time = time_to_value()

		try:
			log = ImportDiscountLog()
			log.status = 1
			log.originalFilename = args['fileName']
			log.resultFilename = filename
			log.operatorUsername = g.current_user.username
			log.timeLower = begin_time
			log.timeUpper = end_time
			log.failcount = count

			db.session.add(log)
			db.session.commit()
		except:
			db.session.rollback()
			db.session.remove()
			abort(500)
		return make_response([])

class ImportDiscount(Resource):

	def post(self):
		from werkzeug.datastructures import FileStorage
		from openpyxl import load_workbook
		import os
		parser = RequestParser(trim=True)
		parser.add_argument('original', type=FileStorage, location='files', required=True)
		args = parser.parse_args(strict=True)
		storage = args['original']

		workbook = load_workbook(storage)
		worksheet = workbook.active
		title = worksheet.delete_rows(0)
		count = 0
		total_balance = 0.0
		total_hitCodeNeed = 0.0
		for row in worksheet.rows:
			try:
				member = Member.query.filter(Member.username == row[0].value).first()
				if not member:
					continue
				if not is_float(row[1].value):
					continue
				if not is_float(row[2].value):
					continue
				count += 1
				total_balance += float(row[1].value)
				total_hitCodeNeed += float(row[2].value)
			except:
				return make_response(error_code=400)

		filename = storage.filename
		worksheet.insert_rows(title)
		workbook.save(os.path.join(current_app.static_folder, filename))

		return make_response([], filename=filename, count=count, total_balance=total_balance, total_hitCodeNeed=total_hitCodeNeed)