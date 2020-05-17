import os, sys, re
import time
import requests
import json
from sqlalchemy import func, distinct, and_
from flask import request, g, current_app
from flask_restful import Resource, marshal_with, fields, abort
from flask_restful.reqparse import RequestParser
import ast
from app.api_0_1.resources.entertainment_city_list import UpdataallWallet
from app.entertainmentcity.amountTransfer import AmountTransfer
from app.models import db
from app.models.common.utils import *
from app.models.entertainment_city import EntertainmentCity
from app.models.entertainment_city_list import YlcOutstationBalance
from app.models.entertainment_city_log import EntertainmentCityTradeLog
from app.models.member import Member, MemberPersonalInfo, MemberLog, MemberAccessLog, MemberBatchCreateLog
from app.models.member_level import MemberLevel
from app.models.blast_bets import BlastBets,BlastBetsCredit
from app.models.bank_account import Bank, MemberBankAccount, MemberBankAccountModificationLog
from app.models.config_fanhui import ConfigFanshui
from app.models.member_account_change import Deposit, Withdrawal
from app.models.message import MessageInbox, MessageOutbox
from app.common.utils import *
from ..common import *
from ..common.utils import *
from app.models.memeber_history import OperationHistory
from sqlalchemy.sql import union,union_all
from sqlalchemy.orm import aliased
from app.models.entertainment_city_bets_detail import EntertainmentCityBetsDetail
from app.common.dataUtils import changeData_str
import os
from openpyxl import Workbook
from flask.json import jsonify


def get_members(criterion, page=None, per_page=None):
	q = db.session.query(
		Member,
		MemberPersonalInfo,
		MemberLevel,
		MemberBankAccount,
		Bank,
		BlastBets,
		MemberAccessLog).group_by(Member.id).order_by(Member.registrationTime.desc())
	q = q.outerjoin(MemberPersonalInfo, Member.id == MemberPersonalInfo.id)
	# q = q.outerjoin(BlastBets, Member.id == BlastBets.uid)
	# q = q.outerjoin(MemberLevel, Member.levelConfig == MemberLevel.id)
	# q = q.outerjoin(MemberBankAccount, Member.id == MemberBankAccount.memberId)
	q = q.outerjoin(Bank, MemberBankAccount.bankId == Bank.id)
	# q = q.outerjoin(MemberAccessLog, Member.id  == MemberAccessLog.memberId)

	q1 = db.session.query(
		MemberLevel.levelName.label('levelName'),
		MemberLevel.id.label('levelId')
	).subquery()
	# q1 = q1.join(MemberLevel, Member.levelConfig == MemberLevel.id).subquery()
	q2 = db.session.query(
		MemberAccessLog.memberId.label('uid'),
		func.max(MemberAccessLog.accessTime).label('accessTime')
	).group_by(MemberAccessLog.memberId).subquery()
	# q2 = q2.join(MemberAccessLog, Member.id == MemberAccessLog.memberId).subquery()

	print(q2)
	print('q2====================================================================')
	q3 = db.session.query(
		BlastBets.uid.label('uid'),
		func.max(BlastBets.actionTime).label('actionTime')
	).group_by(BlastBets.uid).subquery()
	# q2 = q2.join(MemberAccessLog, Member.id == MemberAccessLog.memberId).subquery()
	print(q3)
	print('q3====================================================================')
	# q3 = db.session.query(
	# 	BlastBets.uid.label('uid'),
	# 	func.max(BlastBets.actionTime).label('actionTime')
	# ).group_by(MemberBankAccount.uid).subquery()
	# # q2 = q2.join(MemberAccessLog, Member.id == MemberAccessLog.memberId).subquery()
	# print(q3)
	# print('q3====================================================================')
	q5 = db.session.query(
		Member.id,
		Member.username,
		Member.status,
		Member.balance,
		Member.parent,
		Member.rebateConfig,
		Member.registrationTime,
		Member.type,
		Member.rebate,
		Member.levelConfig,
		Member.rebateRate,
		Member.isTsetPLay,
		q2.c.accessTime.label('accessTime'),
		q3.c.actionTime.label('actionTime'),
		q1.c.levelName.label('levelName'),
		MemberPersonalInfo
	).order_by(Member.registrationTime.desc())
	q5 = q5.outerjoin(q2, Member.id == q2.c.uid)
	q5 = q5.outerjoin(q1, Member.levelConfig == q1.c.levelId)
	q5 = q5.outerjoin(q3, Member.id == q3.c.uid)
	q5 = q5.outerjoin(MemberPersonalInfo, Member.id == MemberPersonalInfo.id)
	print(q5)
	print('q5====================================================================')

	result = []
	total_balance = 0
	total_rebate = 0
	pagination = paginate(q5, criterion, page, per_page)
	print(pagination.items)
	print('================================================================')
	for item in pagination.items:
		print(item.id)
		print(item.username)
		print(item.status)
		print(item.accessTime)
		print('========================================================================')
		total_balance += item.balance or 0
		total_rebate += item.rebate or 0

		parent_username = None
		if item.parent:
			parent = Member.query.get(item.parent)
			if parent:
				parent_username = parent.username

		rebate_config_name = None
		rebate_config = ConfigFanshui.query.get(item.rebateConfig)
		if rebate_config:
			rebate_config_name = rebate_config.name
		result.append({
			'id': item.id,
			'username': item.username,
			'parentUsername': parent_username,
			'name': item.MemberPersonalInfo.name if item.MemberPersonalInfo else None,
			'registrationTime': item.registrationTime,
			'registrationHost': value_to_host(item.registrationTime),
			'status': item.status,
			'type': item.type,
			'balance': item.balance,
			'rebate': item.rebate,
			'rebateConfigName': rebate_config_name,
			'rebateRate': round(((float(item.rebateRate)) * 100), 3),
			'levelConfig': item.levelConfig,
			'levelName': item.levelName,
			'bankName': None,
			'bankAccountName': None,
			'bankAccountNumber': None,
		})
		# print(result)
		# break
		# result.append({
		# 	'id': item.Member.id,
		# 	'username': item.Member.username,
		# 	'parentUsername': parent_username,
		# 	'name': item.MemberPersonalInfo.name if item.MemberPersonalInfo else None,
		# 	'registrationTime': item.Member.registrationTime,
		# 	'registrationHost': value_to_host(item.Member.registrationTime),
		# 	'status': item.Member.status,
		# 	'type': item.Member.type,
		# 	'balance': item.Member.balance,
		# 	'rebate': item.Member.rebate,
		# 	'rebateConfigName': rebate_config_name,
		# 	'rebateRate': round(((float(item.Member.rebateRate))*100),3),
		# 	'levelConfig': item.Member.levelConfig,
		# 	'levelName': item.MemberLevel.levelName if item.MemberLevel else None,
		# 	'bankName': item.Bank.name if item.Bank else None,
		# 	'bankAccountName': item.MemberBankAccount.accountName if item.MemberBankAccount else None,
		# 	'bankAccountNumber': item.MemberBankAccount.accountNumber if item.MemberBankAccount else None,
		# })
	return result, pagination, total_balance, total_rebate


class Members(Resource):

	@marshal_with(make_marshal_fields({
		'id': fields.Integer,
		'username': fields.String,
		'parentUsername': fields.String,
		'name': fields.String,
		'registrationTime': fields.Integer,
		'registrationHost': fields.String,
		'status': fields.Integer,
		'type': fields.Integer,
		'balance': fields.Float,
		'rebate': fields.Float,
		'rebateConfig': fields.Integer,
		'rebateConfigName': fields.String,
		'rebateRate': fields.Float,
		'levelConfig': fields.Integer,
		'levelName': fields.String,
		'bankName': fields.String,
		'bankAccountName': fields.String,
		'bankAccountNumber': fields.String,
		'isBanned': fields.Integer,
		'isSuperMember': fields.Integer
	}, totalBalance=fields.Float, totalRebate=fields.Float))
	def get(self, id=None):
		parser = RequestParser(trim=True)
		parser.add_argument('page', type=int, default=DEFAULT_PAGE)
		parser.add_argument('pageSize', type=int, default=DEFAULT_PAGE_SIZE)

		parser.add_argument('username', type=str)
		parser.add_argument('usernameLike', type=str)
		parser.add_argument('registrationTimeLower', type=int)
		parser.add_argument('registrationTimeUpper', type=int)
		parser.add_argument('balanceLower', type=float)
		parser.add_argument('balanceUpper', type=float)
		parser.add_argument('rebateLower', type=float)
		parser.add_argument('rebateUpper', type=float)
		parser.add_argument('rebateConfig', type=int)
		parser.add_argument('rebateRate', type=float)
		parser.add_argument('status', type=int)
		parser.add_argument('levelId', type=str, dest='levelConfig')
		parser.add_argument('parentUsername', type=str)
		parser.add_argument('wallet', type=str)

		parser.add_argument('name', type=str)
		parser.add_argument('nameLike', type=str)
		parser.add_argument('birthdateLower', type=int)
		parser.add_argument('birthdateUpper', type=int)
		parser.add_argument('gender', type=int)
		parser.add_argument('phone', type=str)
		parser.add_argument('phoneLike', type=str)
		parser.add_argument('email', type=str)
		parser.add_argument('emailLike', type=str)
		parser.add_argument('tencentWeChat', type=str)
		parser.add_argument('tencentWeChatLike', type=str)
		parser.add_argument('tencentQQ', type=str)
		parser.add_argument('tencentQQLike', type=str)

		parser.add_argument('bankAccountNumber', type=str)
		parser.add_argument('bankAccountNumberLike', type=str)
		parser.add_argument('bankAccountName', type=str)
		parser.add_argument('bankAccountNameLike', type=str)

		parser.add_argument('lastLoginTimeLower', type=int)
		parser.add_argument('lastLoginTimeUpper', type=int)

		parser.add_argument('lastBetTimeLower', type=int)
		parser.add_argument('lastBetTimeUpper', type=int)

		parser.add_argument('lastBetTimeLowers', type=int)
		parser.add_argument('lastBetTimeUppers', type=int)
		args = parser.parse_args(strict=True)
		criterion = set()
		criterion.add(Member.type == 0)
		criterion.add(Member.isTsetPLay != 1)
		if id:
			criterion.add(Member.id == id)
		if args['username']:
			criterion.add(Member.username.in_(args['username'].split(',')))
		if args['usernameLike']:
			criterion.add(Member.username.like('%' + args['usernameLike'] + '%'))
		if args['registrationTimeLower']:
			criterion.add(Member.registrationTime >= args['registrationTimeLower'])
		if args['registrationTimeUpper']:
			criterion.add(Member.registrationTime <= args['registrationTimeUpper']  + SECONDS_PER_DAY)
		if args['balanceLower']:
			criterion.add(Member.balance >= args['balanceLower'])
		if args['balanceUpper']:
			criterion.add(Member.balance <= args['balanceUpper'])
		if args['rebateLower']:
			criterion.add(Member.rebate >= args['rebateLower'])
		if args['rebateUpper']:
			criterion.add(Member.rebate <= args['rebateUpper'])
		if args['rebateConfig']:
			criterion.add(Member.rebateConfig == args['rebateConfig'])
		if args['rebateRate'] is not None:
			criterion.add(Member.rebateRate == args['rebateRate'])
		if args['status'] is not None:
			criterion.add(Member.status == args['status'])
		if args['levelConfig']:
			criterion.add(Member.levelConfig.in_(args['levelConfig'].split(',')))
		if args['parentUsername']:
			parent = Member.query.filter(Member.username == args['parentUsername']).first()
			if parent:
				criterion.add(Member.parent == parent.id)
			else:
				return make_response([])
		if args['wallet']:
			pass

		if args['name']:
			criterion.add(MemberPersonalInfo.name == args['name'])
		if args['nameLike']:
			criterion.add(MemberPersonalInfo.name.like('%' + args['nameLike'] + '%'))
		if args['birthdateLower']:
			criterion.add(MemberPersonalInfo.birthdate >= args['birthdateLower'])
		if args['birthdateUpper']:
			criterion.add(MemberPersonalInfo.birthdate <= args['birthdateUpper']  + SECONDS_PER_DAY)
		if args['gender']:
			criterion.add(MemberPersonalInfo.gender == args['gender'])
		if args['phone']:
			criterion.add(MemberPersonalInfo.phone == args['phone'])
		if args['phoneLike']:
			criterion.add(MemberPersonalInfo.phone.like('%' + args['phoneLike'] + '%'))
		if args['email']:
			criterion.add(MemberPersonalInfo.email == args['email'])
		if args['emailLike']:
			criterion.add(MemberPersonalInfo.email.like('%' + args['emailLike'] + '%'))
		if args['tencentWeChat']:
			criterion.add(MemberPersonalInfo.tencentWeChat == args['tencentWeChat'])
		if args['tencentWeChatLike']:
			criterion.add(MemberPersonalInfo.tencentWeChat.like('%' + args['tencentWeChatLike'] + '%'))
		if args['tencentQQ']:
			criterion.add(MemberPersonalInfo.tencentQQ == args['tencentQQ'])
		if args['tencentQQLike']:
			criterion.add(MemberPersonalInfo.tencentQQ.like('%' + args['tencentQQLike'] + '%'))

		if args['bankAccountNumber']:
			str_args = '''select uid from blast_member_bank where account = {}'''.format(args['bankAccountNumber'])
			uidArray = execute(str_args)[0]
			h_results = []
			for numbers in uidArray:
				h_results.append(numbers[0])
			criterion.add(Member.id.in_(h_results))
		if args['bankAccountNumberLike']:
			str_args = '''select uid from blast_member_bank where account like "%{}%"'''.format(args["bankAccountNumberLike"])
			uidArray = execute(str_args)[0]
			h_results = []
			for numbers in uidArray:
				h_results.append(numbers[0])
			criterion.add(Member.id.in_(h_results))
		if args['bankAccountName']:
			criterion.add(MemberBankAccount.accountName == args['bankAccountName'])
		if args['bankAccountNameLike']:
			criterion.add(MemberBankAccount.accountName.like('%' + args['bankAccountNameLike'] + '%'))

		if not args['pageSize']:
			args['pageSize'] = Member.query.count()

		criterion_q2 = set()
		if args['lastLoginTimeLower']:
			criterion_q2.add(MemberAccessLog.accessTime >= args['lastLoginTimeLower'])
		if args['lastLoginTimeUpper']:
			criterion_q2.add(MemberAccessLog.accessTime <= args['lastLoginTimeUpper']  + SECONDS_PER_DAY)
		if args['lastLoginTimeUpper'] or args['lastLoginTimeLower']:
			q2 = db.session.query(
				MemberAccessLog.memberId.label('uid')
			).filter(*criterion_q2)
			criterion.add(Member.id.in_(q2))

		criterion_q3 = set()
		criterion_q4 = set()
		criterion_ci = set()
		if args['lastBetTimeLower']:
			criterion_q3.add(BlastBets.actionTime >= args['lastBetTimeLower'])
			criterion_q4.add(BlastBetsCredit.betTime >= args['lastBetTimeLower'])
			criterion_ci.add(EntertainmentCityBetsDetail.BetTime >= args['lastBetTimeLower'])
		if args['lastBetTimeUpper']:
			criterion_q3.add(BlastBets.actionTime <= args['lastBetTimeUpper'])
			criterion_q4.add(BlastBetsCredit.betTime <= args['lastBetTimeUpper'])
			criterion_ci.add(EntertainmentCityBetsDetail.BetTime <= args['lastBetTimeUpper'])
		if args['lastBetTimeUpper'] or args['lastBetTimeLower']:
			betsAccount_guan = db.session.query(
				BlastBets.username.label('username')
			).filter(*criterion_q3)
			betsAccount_xinyong = db.session.query(
				BlastBetsCredit.memberUsername.label('username')
			).filter(*criterion_q4)
			betsAccount_city = db.session.query(
				EntertainmentCityBetsDetail.PlayerName.label('username')
			).filter(*criterion_ci)
			result = union_all(betsAccount_guan, betsAccount_xinyong,betsAccount_city)
			user_alias = aliased(result, name='user_alias')
			q3 = db.session.query(
				user_alias.c.username.label('username')
			).distinct()
			criterion.add(Member.username.in_(q3))

		q1 = db.session.query(
			MemberLevel.levelName.label('levelName'),
			MemberLevel.id.label('levelId')
		).subquery()
		q5 = db.session.query(
			Member.id,
			Member.username,
			Member.status,
			Member.balance,
			Member.parent,
			Member.rebateConfig,
			Member.registrationTime,
			Member.type,
			Member.rebate,
			Member.levelConfig,
			Member.rebateRate,
			Member.isTsetPLay,
			Member.isBanned,
			Member.isSuperMember,

			q1.c.levelName.label('levelName'),
			MemberPersonalInfo
		).order_by(Member.registrationTime.desc())
		q5 = q5.outerjoin(q1, Member.levelConfig == q1.c.levelId)
		q5 = q5.outerjoin(MemberPersonalInfo, Member.id == MemberPersonalInfo.id)

		# if args['lastLoginTimeLower']:
		# 	criterion.add(q2.c.accessTime >= args['lastLoginTimeLower'])
		# if args['lastLoginTimeUpper']:
		# 	criterion.add(q2.c.accessTime <= args['lastLoginTimeUpper'])
		#
		# if args['lastBetTimeLower']:
		# 	criterion.add(q3.c.actionTime >= args['lastBetTimeLower'])
		# if args['lastBetTimeUpper']:
		# 	criterion.add(q3.c.actionTime <= args['lastBetTimeUpper'])
		# jsonlist = db.session.query(YlcOutstationBalance.uid, YlcOutstationBalance.jsonlist).all()
		# znye_mem = []
		# if jsonlist:
		# 	for jslt in jsonlist:
		# 		znye = {}
		# 		zn = eval(''.join(jslt[1]))
		# 		balanceOuter = 0
		# 		for z in zn:
		# 			a = float(z['balance'])
		# 			balanceOuter += a
		# 		znye['id'] = jslt[0]
		# 		znye['balanceOuter'] = balanceOuter
		# 		znye_mem.append(znye)

		result = []
		total_balance = 0
		total_rebate = 0
		pagination = paginate(q5, criterion, args['page'], args['pageSize'])
		for item in pagination.items:
			total_balance += item.balance or 0
			total_rebate += item.rebate or 0

			parent_username = None
			if item.parent:
				parent = Member.query.get(item.parent)
				if parent:
					parent_username = parent.username

			rebate_config_name = None
			if item.rebateConfig is not None:
				rebate_config = ConfigFanshui.query.get(item.rebateConfig)
			else:
				rebate_config = None
			if rebate_config:
				rebate_config_name = rebate_config.name
			balanceOuter = 0
			# for outer in znye_mem:
			# 	if item.id == outer['id']:
			# 		balanceOuter = outer['balanceOuter']
			result.append({
				'id': item.id,
				'username': item.username,
				'parentUsername': parent_username,
				'name': item.MemberPersonalInfo.name if item.MemberPersonalInfo else None,
				'registrationTime': item.registrationTime,
				'registrationHost': value_to_host(item.registrationTime),
				'status': item.status,
				'type': item.type,
				'balance': item.balance + balanceOuter,
				'rebate': item.rebate,
				'rebateConfig': item.rebateConfig,
				'rebateConfigName': rebate_config_name,
				'rebateRate': item.rebateRate,
				'levelConfig': item.levelConfig,
				'levelName': item.levelName,
				'bankName': None,
				'bankAccountName': None,
				'bankAccountNumber': None,
				'isBanned': item.isBanned,
				'isSuperMember': item.isSuperMember,
			})
		return make_response(result,
			page=pagination.page, pages=pagination.pages, total=pagination.total,
			totalBalance=args['page'], totalRebate=args['pageSize']
		)

	def post(self):
		member_parser = RequestParser(trim=True)
		member_parser.add_argument('username', type=str, required=True)
		member_parser.add_argument('password', type=str, default=DEFAULT_MEMBER_PASSWORD)
		member_parser.add_argument('fundPassword', type=str, default=DEFAULT_MEMBER_FUND_PASSWORD)
		member_parser.add_argument('parentUsername', type=str, required=True)
		member_parser.add_argument('rebateRate', type=float, required=True)
		member_parser.add_argument('name', type=str, required=True)
		member_parser.add_argument('isBanned', type=int)
		member_parser.add_argument('isSuperMember', type=int)
		member_args = member_parser.parse_args()
		member_args = {key: value for key, value in member_args.items() if value is not None}

		member_personal_info_parser = RequestParser(trim=True)
		member_personal_info_parser.add_argument('name', type=str)
		member_personal_info_parser.add_argument('birthdate', type=int)
		member_personal_info_parser.add_argument('gender', type=int)
		member_personal_info_parser.add_argument('phone', type=str)
		member_personal_info_parser.add_argument('email', type=str)
		member_personal_info_parser.add_argument('tencentQQ', type=str)
		member_personal_info_parser.add_argument('tencentWeChat', type=str)
		member_personal_info_parser.add_argument('remark', type=str)
		member_personal_info_args = member_personal_info_parser.parse_args()
		member_personal_info_args = {key: value for key, value in member_personal_info_args.items() if value}
# 		if 'rebateRate' in member_args and member_args['rebateRate']:
# 			member_args['rebateRate'] = float(float(member_args['rebateRate'])/100)
		pattern = "^[A-Za-z0-9\\u4e00-\\u9fa5][A-Za-z0-9\\u4e00-\\u9fa5-_]{4,13}[A-Za-z0-9\\u4e00-\\u9fa5]$"
		re_result = re.match(pattern, member_args['username'])
		if re_result is None:
			return make_response(error_code=400, error_message="账号必须为6-15位字母和数字组合")
		member = Member.query.filter(Member.username == member_args['username']).first()
		if member:
			return make_response(error_code=400, error_message="该用户已存在,请重新输入")
		parent = Member.query.filter(Member.username == member_args.pop('parentUsername'),Member.type==1).first()
		if not parent:
			return make_response(error_code=400, error_message="该代理不存在,请重新输入")
		if parent.status == 0:
			return jsonify({
				'success': False,
				'errorCode': 403,
				'errorMsg': '该代理已被停用'
			})

		if member_args['rebateRate'] is not None:
			m_boolean = member_args['rebateRate'] <= parent.rebateRate or member_args['rebateRate']==0
			if not m_boolean:
				return make_response(error_code=400, error_message="返点率不能超过"+str(float(parent.rebateRate))+"%")
		else:
			return make_response(error_code=400, error_message="返点不正确")
		parents = parent.parents + ',%s'%parent.id
		parentsInfo = parent.parentsInfo + ',%s'%parent.username
		try:
			member_args['type'] = 0
			member_args['parent'] = parent.id
			member_args['parents'] = parents
			member_args['parentsInfo'] = parentsInfo
			member_args['levelConfig'] = parent.defaultLevelConfig
			member_args['rebateConfig'] = parent.defaultRebateConfig
			member_args['defaultLevelConfig'] = parent.defaultLevelConfig
			member_args['defaultRebateConfig'] = parent.defaultRebateConfig
			member_args['commissionConfig'] = parent.commissionConfig
			member_args['registrationTime'] = time_to_value()
			member_args['registrationHost'] = host_to_value(request.remote_addr)
			member_args['createrUsername'] = g.current_user.username
			member = Member(**member_args)
			db.session.add(member)
			db.session.commit()
			OperationHistory().PublicMemberDatasApply(1001, member.id)
			try:
				member_personal_info_args['id'] = member.id
				member_personal_info = MemberPersonalInfo(**member_personal_info_args)
				db.session.add(member_personal_info)
				db.session.commit()
				OperationHistory().PublicMemberDatasApply(1002, member.id)
			except:
				db.session.rollback()
				db.session.remove()
				db.session.delete(member)
				db.session.commit()
				abort(500)
		except:
			db.session.rollback()
			db.session.remove()
			abort(500)
		return make_response([]), 201

	def put(self):
		parser = RequestParser(trim=True)
		parser.add_argument('memberId', type=str, required=True)
		parser.add_argument('status', type=int)
		parser.add_argument('levelId', type=int, dest='levelConfig')
		parser.add_argument('rebateRate', type=float)
		parser.add_argument('rebate', type=float)
		parser.add_argument('rebateConfig', type=int)
		parser.add_argument('parentId', type=int)

		parser.add_argument('isBanned', type=int)
		parser.add_argument('isSuperMember', type=int)
		args = parser.parse_args(strict=True)
		args = {key: value for key, value in args.items() if value is not None}
		if 'rebateRate' in args and args['rebateRate'] is not None:
			if 'parentId' not in args:
				return make_response(error_code=400, error_message="找不到上级")
			parent = Member.query.filter(Member.id == args.pop('parentId')).first()
			if not parent or not parent.type:
				return make_response(error_code=400, error_message="该代理不存在,请重新输入")
			m_boolean = args['rebateRate'] <= parent.rebateRate or args['rebateRate']==0
			if not m_boolean:
				return make_response(error_code=400, error_message="返点率不能超过" + str(float(parent.rebateRate)) + "%")
		if 'parentId' in args:
			args.pop('parentId')
		try:
			member_id_list = args.pop('memberId').split(',')
			Member.update_all(Member.get_all(member_id_list), args)
		except:
			db.session.rollback()
			db.session.remove()
			abort(500)
		return make_response([])


class ResetPassword(Resource):

	@marshal_with(make_marshal_fields({
		'password': fields.String
	}))
	def put(self):
		parser = RequestParser(trim=True)
		parser.add_argument('username', type=str, required=True)
		args = parser.parse_args(strict=True)

		password = Member.reset_password(args['username'])
		return make_response([{
			'password': password
		}])


class ResetFundPassword(Resource):

	@marshal_with(make_marshal_fields({
		'password': fields.String
	}))
	def put(self):
		parser = RequestParser(trim=True)
		parser.add_argument('username', type=str, required=True)
		args = parser.parse_args(strict=True)

		password = Member.reset_fund_password(args['username'])
		return make_response([{
			'password': password
		}])


class ExportMembers(Resource):
	def get(self):
		parser = RequestParser(trim=True)
		parser.add_argument('username', type=str)
		parser.add_argument('usernameLike', type=str)
		parser.add_argument('registrationTimeLower', type=int)
		parser.add_argument('registrationTimeUpper', type=int)
		parser.add_argument('balanceLower', type=float)
		parser.add_argument('balanceUpper', type=float)
		parser.add_argument('rebateLower', type=float)
		parser.add_argument('rebateUpper', type=float)
		parser.add_argument('rebateConfig', type=int)
		parser.add_argument('rebateRate', type=float)
		parser.add_argument('status', type=int)
		parser.add_argument('levelId', type=str, dest='levelConfig')
		parser.add_argument('parentUsername', type=str)
		parser.add_argument('wallet', type=str)

		parser.add_argument('name', type=str)
		parser.add_argument('nameLike', type=str)
		parser.add_argument('birthdateLower', type=int)
		parser.add_argument('birthdateUpper', type=int)
		parser.add_argument('gender', type=int)
		parser.add_argument('phone', type=str)
		parser.add_argument('phoneLike', type=str)
		parser.add_argument('email', type=str)
		parser.add_argument('emailLike', type=str)
		parser.add_argument('tencentWeChat', type=str)
		parser.add_argument('tencentWeChatLike', type=str)
		parser.add_argument('tencentQQ', type=str)
		parser.add_argument('tencentQQLike', type=str)

		parser.add_argument('bankAccountNumber', type=str)
		parser.add_argument('bankAccountNumberLike', type=str)
		parser.add_argument('bankAccountName', type=str)
		parser.add_argument('bankAccountNameLike', type=str)

		parser.add_argument('lastLoginTimeLower', type=int)
		parser.add_argument('lastLoginTimeUpper', type=int)

		parser.add_argument('lastBetTimeLower', type=int)
		parser.add_argument('lastBetTimeUpper', type=int)

		parser.add_argument('lastBetTimeLowers', type=int)
		parser.add_argument('lastBetTimeUppers', type=int)
		args = parser.parse_args(strict=True)

		criterion = set()
		criterion.add(Member.type == 0)
		criterion.add(Member.isTsetPLay != 1)
		if args['username']:
			criterion.add(Member.username.in_(args['username'].split(',')))
		if args['usernameLike']:
			criterion.add(Member.username.like('%' + args['usernameLike'] + '%'))
		if args['registrationTimeLower']:
			criterion.add(Member.registrationTime >= args['registrationTimeLower'])
		if args['registrationTimeUpper']:
			criterion.add(Member.registrationTime <= args['registrationTimeUpper']  + SECONDS_PER_DAY)
		if args['balanceLower']:
			criterion.add(Member.balance >= args['balanceLower'])
		if args['balanceUpper']:
			criterion.add(Member.balance <= args['balanceUpper'])
		if args['rebateLower']:
			criterion.add(Member.rebate >= args['rebateLower'])
		if args['rebateUpper']:
			criterion.add(Member.rebate <= args['rebateUpper'])
		if args['rebateConfig']:
			criterion.add(Member.rebateConfig == args['rebateConfig'])
		if args['rebateRate'] is not None:
			criterion.add(Member.rebateRate == args['rebateRate'])
		if args['status'] is not None:
			criterion.add(Member.status == args['status'])
		if args['levelConfig']:
			criterion.add(Member.levelConfig.in_(args['levelConfig'].split(',')))
		if args['parentUsername']:
			parent = Member.query.filter(Member.username == args['parentUsername']).first()
			if parent:
				criterion.add(Member.parent == parent.id)
			else:
				return make_response([])
		if args['wallet']:
			pass

		if args['name']:
			criterion.add(MemberPersonalInfo.name == args['name'])
		if args['nameLike']:
			criterion.add(MemberPersonalInfo.name.like('%' + args['nameLike'] + '%'))
		if args['birthdateLower']:
			criterion.add(MemberPersonalInfo.birthdate >= args['birthdateLower'])
		if args['birthdateUpper']:
			criterion.add(MemberPersonalInfo.birthdate <= args['birthdateUpper']  + SECONDS_PER_DAY)
		if args['gender']:
			criterion.add(MemberPersonalInfo.gender == args['gender'])
		if args['phone']:
			criterion.add(MemberPersonalInfo.phone == args['phone'])
		if args['phoneLike']:
			criterion.add(MemberPersonalInfo.phone.like('%' + args['phoneLike'] + '%'))
		if args['email']:
			criterion.add(MemberPersonalInfo.email == args['email'])
		if args['emailLike']:
			criterion.add(MemberPersonalInfo.email.like('%' + args['emailLike'] + '%'))
		if args['tencentWeChat']:
			criterion.add(MemberPersonalInfo.tencentWeChat == args['tencentWeChat'])
		if args['tencentWeChatLike']:
			criterion.add(MemberPersonalInfo.tencentWeChat.like('%' + args['tencentWeChatLike'] + '%'))
		if args['tencentQQ']:
			criterion.add(MemberPersonalInfo.tencentQQ == args['tencentQQ'])
		if args['tencentQQLike']:
			criterion.add(MemberPersonalInfo.tencentQQ.like('%' + args['tencentQQLike'] + '%'))

		if args['bankAccountNumber']:
			str_args = '''select uid from blast_member_bank where account = {}'''.format(args['bankAccountNumber'])
			uidArray = execute(str_args)[0]
			h_results = []
			for numbers in uidArray:
				h_results.append(numbers[0])
			criterion.add(Member.id.in_(h_results))
		if args['bankAccountNumberLike']:
			str_args = '''select uid from blast_member_bank where account like "%{}%"'''.format(args["bankAccountNumberLike"])
			uidArray = execute(str_args)[0]
			h_results = []
			for numbers in uidArray:
				h_results.append(numbers[0])
			criterion.add(Member.id.in_(h_results))
		if args['bankAccountName']:
			criterion.add(MemberBankAccount.accountName == args['bankAccountName'])
		if args['bankAccountNameLike']:
			criterion.add(MemberBankAccount.accountName.like('%' + args['bankAccountNameLike'] + '%'))

		criterion_q2 = set()
		if args['lastLoginTimeLower']:
			criterion_q2.add(MemberAccessLog.accessTime >= args['lastLoginTimeLower'])
		if args['lastLoginTimeUpper']:
			criterion_q2.add(MemberAccessLog.accessTime <= args['lastLoginTimeUpper']  + SECONDS_PER_DAY)
		if args['lastLoginTimeUpper'] or args['lastLoginTimeLower']:
			q2 = db.session.query(
				MemberAccessLog.memberId.label('uid')
			).filter(*criterion_q2)
			criterion.add(Member.id.in_(q2))

		criterion_q3 = set()
		criterion_q4 = set()
		criterion_ci = set()
		if args['lastBetTimeLower']:
			criterion_q3.add(BlastBets.actionTime >= args['lastBetTimeLower'])
			criterion_q4.add(BlastBetsCredit.betTime >= args['lastBetTimeLower'])
			criterion_ci.add(EntertainmentCityBetsDetail.BetTime >= args['lastBetTimeLower'])
		if args['lastBetTimeUpper']:
			criterion_q3.add(BlastBets.actionTime <= args['lastBetTimeUpper'])
			criterion_q4.add(BlastBetsCredit.betTime <= args['lastBetTimeUpper'])
			criterion_ci.add(EntertainmentCityBetsDetail.BetTime <= args['lastBetTimeUpper'])
		if args['lastBetTimeUpper'] or args['lastBetTimeLower']:
			betsAccount_guan = db.session.query(
				BlastBets.username.label('username')
			).filter(*criterion_q3)
			betsAccount_xinyong = db.session.query(
				BlastBetsCredit.memberUsername.label('username')
			).filter(*criterion_q4)
			betsAccount_city = db.session.query(
				EntertainmentCityBetsDetail.PlayerName.label('username')
			).filter(*criterion_ci)
			result = union_all(betsAccount_guan, betsAccount_xinyong,betsAccount_city)
			user_alias = aliased(result, name='user_alias')
			q3 = db.session.query(
				user_alias.c.username.label('username')
			).distinct()
			criterion.add(Member.username.in_(q3))

		q5 = db.session.query(Member, MemberPersonalInfo, MemberBankAccount, MemberLevel).order_by(
			Member.registrationTime.desc())
		q5 = q5.outerjoin(MemberPersonalInfo, MemberPersonalInfo.id == Member.id)
		q5 = q5.outerjoin(MemberBankAccount, Member.id == MemberBankAccount.memberId)
		q5 = q5.outerjoin(MemberLevel, MemberLevel.id == Member.levelConfig)

		items = q5.filter(*criterion).all()

		jsonlist = db.session.query(YlcOutstationBalance.uid, YlcOutstationBalance.jsonlist).all()
		znye_mem = []
		if jsonlist:
			for jslt in jsonlist:
				znye = {}
				zn = eval(''.join(jslt[1]))
				balanceOuter = 0
				for z in zn:
					a = float(z['balance'])
					balanceOuter += a
				znye['id'] = jslt[0]
				znye['balanceOuter'] = balanceOuter
				znye_mem.append(znye)
		deposit_online_mem = db.session.query(
			Deposit.memberId,
			func.count(Deposit.memberId),
			func.sum(Deposit.applicationAmount),
		).filter(Deposit.type.in_([100003,100004]), Deposit.isAcdemen==1, Deposit.status == 2).group_by(Deposit.memberId).all()

		withdrawal_online_mem = db.session.query(
			Withdrawal.memberId,
			func.count(Withdrawal.memberId),
			func.sum(Withdrawal.applicationAmount),
		).filter(Withdrawal.type == 200001, Withdrawal.isAcdemen == 1, Withdrawal.status == 2).group_by(
			Withdrawal.memberId).all()

		deposit_sjct_mem = db.session.query(
			Deposit.memberId,
			func.count(Deposit.memberId),
			func.sum(Deposit.applicationAmount),
		).filter(Deposit.type==900001, Deposit.isAcdemen==1, Deposit.status == 2).group_by(Deposit.memberId).all()

		withdrawal_sjct_mem = db.session.query(
			Withdrawal.memberId,
			func.count(Withdrawal.memberId),
			func.sum(Withdrawal.applicationAmount),
		).filter(Withdrawal.type == 900002, Withdrawal.isAcdemen == 1, Withdrawal.status == 2).group_by(Withdrawal.memberId).all()

		results = []
		for item in items:
			member_bank = Bank.query.get(item.MemberBankAccount.bankId) if item.MemberBankAccount else None
			balanceOuter = 0
			for outer in znye_mem:
				if item.Member.id == outer['id']:
					balanceOuter = outer['balanceOuter']
			deposit_online_count = 0
			deposit_online_amount = 0.0
			if deposit_online_mem:
				for deposit_online in deposit_online_mem:
					if item.Member.id == deposit_online[0]:
						deposit_online_count = deposit_online[1]
						deposit_online_amount = deposit_online[2]

			withdrawal_online_count = 0
			withdrawal_online_amount = 0.0
			if withdrawal_online_mem:
				for withdrawal_online in withdrawal_online_mem:
					if item.Member.id == withdrawal_online[0]:
						withdrawal_online_count = withdrawal_online[1]
						withdrawal_online_amount = withdrawal_online[2]

			deposit_sjct_count = 0
			deposit_sjct_amount = 0.0
			if deposit_sjct_mem:
				for deposit_sjct in deposit_sjct_mem:
					if item.Member.id == deposit_sjct[0]:
						deposit_sjct_count = deposit_sjct[1]
						deposit_sjct_amount = deposit_sjct[2]

			withdrawal_sjct_count = 0
			withdrawal_sjct_amount = 0.0
			if deposit_online_mem:
				for withdrawal_sjct in withdrawal_sjct_mem:
					if item.Member.id == withdrawal_sjct[0]:
						withdrawal_sjct_count = withdrawal_sjct[1]
						withdrawal_sjct_amount = withdrawal_sjct[2]

			results.append([
				item.Member.username,
				item.MemberLevel.levelName if item.MemberLevel else None,
				item.MemberPersonalInfo.name if item.MemberPersonalInfo else None,
				item.MemberBankAccount.accountNumber if item.MemberBankAccount else None,
				item.Member.registrationTime,
				item.Member.status,
				item.Member.balance + balanceOuter,
				item.Member.balance,
				balanceOuter,
				0,
				float(item.Member.rebateRate),
				deposit_online_count,
				deposit_online_amount,
				withdrawal_online_count,
				withdrawal_online_amount,
				deposit_sjct_count,
				deposit_sjct_amount,
				withdrawal_sjct_count,
				withdrawal_sjct_amount,
			])

		from openpyxl import Workbook
		workbook = Workbook()
		worksheet = workbook.active
		title = ['账号', '会员等级', '真实姓名', '银行账户', '入会日期', '状态', '账户余额', '站内余额',
				 '站外余额', '优惠钱包', '返点率', '在线存款次数', '在线存款金额', '在线取款次数', '在线取款金额',
				 '人工存款(实际存提)次数', '人工存款(实际存提)金额', '人工取款(实际存提)次数', '人工取款(实际存提)金额']
		worksheet.append(title)
		for result in results:
			st1 = time.localtime(result[4])
			t1 = time.strftime('%Y-%m-%d %H:%M:%S', st1)
			result[4] = t1
			if result[5] == 0:
				result[5] = '禁用'
			elif result[5] == 1:
				result[5] = '启用'
			elif result[5] == 2:
				result[5] = '钱包冻结'
			worksheet.append(result)
		filename = '会员-' + str(int(time.time())) + '.xlsx'
		workbook.save(os.path.join(current_app.static_folder, filename))

		return make_response([{
			'success': True,
			'resultFilename': filename,
		}])


class ImportMembers(Resource):

	@marshal_with(make_marshal_fields({
		'id': fields.Integer,
		'username': fields.String
	}, resultFilename=fields.String, count=fields.Integer))
	def post(self):
		from werkzeug.datastructures import FileStorage
		parser = RequestParser(trim=True)
		parser.add_argument('original', type=FileStorage, location='files', required=True)
		args = parser.parse_args(strict=True)

		from openpyxl import load_workbook, Workbook
		workbook = load_workbook(args['original'])
		worksheet = workbook.active
		worksheet.delete_rows(0)
		try:
			usernames = []
			for row in worksheet.rows:
				if row[0]:
					usernames.append(row[0].value)
			m_res = Member.query.filter(Member.username.in_(usernames),Member.type == 0).all()
			count = len(usernames)
			result = []
			for res in m_res:
				result.append(res.username)
			workBook = Workbook()
			workSheet = workBook.active
			title = ['账号', '错误讯息']
			workSheet.append(title)
			for name in usernames:
				m_result = []
				if name in result:
					count -= 1
					m_result.append(name)
					workSheet.append(m_result)
				else:
					m_result.append(name)
					m_result.append('查无此帐号')
					workSheet.append(m_result)
			filename = '汇入大量帐号-' + str(int(time.time())) + '.xlsx'
			workBook.save(os.path.join(current_app.static_folder, filename))
			return make_response(
				data = m_res,
				resultFilename = filename,
				count = count,
			)
		except:
			return make_response(error_code=400)


class MemberPersonalInfos(Resource):

	@marshal_with(make_marshal_fields({
		'id': fields.Integer,
		'name': fields.String,
		'gender': fields.Integer,
		'birthdate': fields.Integer,
		'phone': fields.String,
		'email': fields.String,
		'tencentQQ': fields.String,
		'tencentWeChat': fields.String,
		'remark': fields.String,
	}))
	def get(self, id):
		member = Member.query.get(id)
		if not member:
			abort(400)
		return make_response(member.personalInfo.all())

	def put(self, id):
		parser = RequestParser(trim=True)
		parser.add_argument('name', type=str)
		parser.add_argument('gender', type=int)
		parser.add_argument('birthdate', type=int)
		parser.add_argument('phone', type=str)
		parser.add_argument('email', type=str)
		parser.add_argument('tencentQQ', type=str)
		parser.add_argument('tencentWeChat', type=str)
		args = parser.parse_args(strict=True)
		member = Member.query.get(id)
		if not member:
			abort(400)
		try:
			memberpersonalInfo = MemberPersonalInfo.query.filter(MemberPersonalInfo.id == member.id).all()
			if memberpersonalInfo:
				member.personalInfo.update(args)
			else:
				args['id'] = member.id
				dao = MemberPersonalInfo(**args)
				db.session.add(dao)
			db.session.commit()
			OperationHistory().PublicMeDatas(2222,id)
		except:
			db.session.rollback()
			db.session.remove()
			abort(500)
		return make_response([])
class MemberPersonalInfosRemark(Resource):
	def put(self,id):
		parser = RequestParser(trim=True)
		parser.add_argument('remark', type=str)
		args = parser.parse_args(strict=True)
		# args = {key: value for key, value in args.items() if value}
		member = Member.query.get(id)
		if not member:
			abort(400)
		try:
			memberpersonalInfo = MemberPersonalInfo.query.filter(MemberPersonalInfo.id == member.id).all()
			if memberpersonalInfo:
				member.personalInfo.update(args)
			else:
				args['id'] = member.id
				dao = MemberPersonalInfo(**args)
				db.session.add(dao)
			db.session.commit()
		except:
			db.session.rollback()
			db.session.remove()
			abort(500)
		return make_response([])


class MemberDetails(Resource):

	@marshal_with(make_marshal_fields({
		'id': fields.Integer,
		'username': fields.String,
		'registrationTime': fields.Integer,
		'registrationHost': fields.String,
		'status': fields.Integer,
		'levelName': fields.String,
		'levelConfig': fields.Integer,
		'parents': fields.List(fields.Nested({
			'id': fields.Integer,
			'username': fields.String,
			'rebateRate':fields.Float,
		})),
		'balance': fields.Float,
		'balanceInner': fields.Float,
		'balanceOuter': fields.Float,
		'rebateRate': fields.Float,
		'rebate': fields.Float,
		'defaultRebateConfig': fields.Integer,
		'rebateConfigName': fields.String,
		'rebateConfig': fields.Integer,
		'name': fields.String,
		'phone': fields.String,
		'remark': fields.String,
		'bankId': fields.Integer,
		'bankName': fields.String,
		'bankAccountName': fields.String,
		'bankAccountNumber': fields.String,
		'totalDepositTimes': fields.Integer,
		'totalDepositAmount': fields.Float,
		'totalWithdrawalTimes': fields.Integer,
		'totalWithdrawalAmount': fields.Float,
		'lastLoginTime': fields.Integer,
		'lastLoginHost': fields.String,
		'numberOfMessage': fields.Integer,
		'isSuperMember': fields.Integer,
		'isBanned': fields.Integer,
		'numberOfUnreadMessage': fields.Integer
	}))
	def get(self, id):
		member = Member.query.get(id)
		if not member:
			return make_response([])

		# 站内余额已处理
		# 站外余额已处理
		jsonlist = db.session.query(YlcOutstationBalance.jsonlist).filter(
			YlcOutstationBalance.username == member.username).first()

		balanceOuter = 0
		if jsonlist:
			jsonlist = ast.literal_eval(''.join(jsonlist))

			for jslt in jsonlist:
				a = float(jslt['balance'])
				print(a)
				balanceOuter += a


		result = {
			'id': member.id,
			'username': member.username,
			'isBanned': member.isBanned,
			'isSuperMember': member.isSuperMember,
			'registrationTime': member.registrationTime,
			'registrationHost': value_to_host(member.registrationHost),
			'status': member.status,
			'levelConfig': member.levelConfig if member.levelConfig else None,
			'parents': [],
			'balanceInner': member.balance,
			'balanceOuter': balanceOuter,
			'balance': member.balance,
			'rebateRate':member.rebateRate,
			'rebate': member.rebate,
			'defaultRebateConfig': member.defaultRebateConfig,
			'rebateConfigName': member.rebate_config.name if member.rebate_config else None,
			'rebateConfig': member.rebateConfig
		}

		member_level = MemberLevel.query.get(member.levelConfig)
		if member_level:
			result['levelName'] = member_level.levelName

		parent_id = member.parent
		while parent_id:
			parent = Member.query.get(parent_id)
			#太尼玛危险了，很可能造成内存泄漏，死循环
			if parent_id == parent.parent:
				break
			if not parent:
				break
			result['parents'].append({'id': parent.id, 'username': parent.username,"rebateRate":parent.rebateRate})
			parent_id = parent.parent
			
		personal_info = member.personalInfo.first()
		if personal_info:
			result['name'] = personal_info.name
			result['phone'] = personal_info.phone
			result['remark'] = personal_info.remark

		bank_account = member.bankAccount.first()
		if bank_account:
			result['bankId'] = bank_account.bankId
			result['bankName'] = bank_account.bank.name
			result['bankAccountName'] = bank_account.accountName
			result['bankAccountNumber'] = bank_account.accountNumber


		result_deposits = db.session.query(func.count(Deposit.memberId), func.sum(Deposit.applicationAmount)).filter(
			Deposit.memberId == id,Deposit.type.in_([100003,100004]),
			Deposit.status == 2
		).all()[0]
		result['totalDepositTimes'] = result_deposits[0]
		result['totalDepositAmount'] = result_deposits[1]

		result_withdrawals = db.session.query(func.count(Withdrawal.memberId), func.sum(Withdrawal.withdrawalAmount)).filter(
			Withdrawal.memberId == id,Withdrawal.type == 200001,Withdrawal.status == 2).all()[0]
		result['totalWithdrawalTimes'] = result_withdrawals[0]
		result['totalWithdrawalAmount'] = result_withdrawals[1]

		MessageInbox_count = db.session.query(func.count(MessageInbox.senderid)).filter(MessageInbox.senderid == id).all()[0]
		MessageInbox_read = db.session.query(func.count(MessageInbox.senderid)).filter(MessageInbox.senderid == id, MessageInbox.read == 0).all()[0]
		result['numberOfMessage'] = MessageInbox_count[0]
		result['numberOfUnreadMessage'] = MessageInbox_read[0]

		result_memberAccessLog = db.session.query(MemberAccessLog.accessTime, MemberAccessLog.realIP).filter(
			MemberAccessLog.memberId == id).order_by(MemberAccessLog.accessTime.desc()).first()
		if result_memberAccessLog:
			result['lastLoginTime'] = result_memberAccessLog[0]
			result['lastLoginHost'] = result_memberAccessLog[1]

		return make_response([result])


class MemberAccessLogs(Resource):

	@marshal_with(make_marshal_fields({
		'id': fields.Integer,
		'memberId': fields.Integer,
		'username': fields.String,
		'accessTime': fields.Integer,
		'realIP': fields.String,
		'accessWebsite': fields.String,
		'country': fields.String,
		'province': fields.String,
		'city': fields.String,
		'operator': fields.String,
	}))
	def get(self, id=None):
		parser = RequestParser(trim=True)
		parser.add_argument('page', type=int, default=DEFAULT_PAGE)
		parser.add_argument('pageSize', type=int, default=DEFAULT_PAGE_SIZE)
		parser.add_argument('username', type=str)
		parser.add_argument('accessTimeLower', type=int)
		parser.add_argument('accessTimeUpper', type=int)
		parser.add_argument('realIP', type=str)
		parser.add_argument('realIPLike', type=str)
		parser.add_argument('accessWebsite', type=str)
		parser.add_argument('accessWebsiteLike', type=str)
		parser.add_argument('isLast', type=int)
		args = parser.parse_args(strict=True)

		criterion = set()
		if id:
			criterion.add(MemberAccessLog.id.in_([id]))
		if args['username']:
			criterion.add(MemberAccessLog.username == args['username'])
		if args['accessTimeLower']:
			criterion.add(MemberAccessLog.accessTime >= args['accessTimeLower'])
		if args['accessTimeUpper']:
			criterion.add(MemberAccessLog.accessTime <= args['accessTimeUpper'] + SECONDS_PER_DAY)
		if args['realIP']:
			criterion.add(MemberAccessLog.realIP == args['realIP'])
		if args['realIPLike']:
			criterion.add(MemberAccessLog.realIP.like('%' + args['realIPLike'] + '%'))
		if args['accessWebsite']:
			criterion.add(MemberAccessLog.accessWebsite == args['accessWebsite'])
		if args['accessWebsiteLike']:
			criterion.add(MemberAccessLog.accessWebsite.like('%' + args['accessWebsiteLike'] + '%'))
		if args['isLast'] == 1:
			str_args = '''SELECT * FROM blast_member_session a,
						  (SELECT uid, MAX(loginTime) AS loginTime FROM blast_member_session GROUP BY uid) b
						  WHERE a.uid = b.uid AND a.loginTime = b.loginTime;'''
			uidArray = execute(str_args)[0]
			h_results = []
			for numbers in uidArray:
				h_results.append(numbers[0])
			criterion.add(MemberAccessLog.id.in_(h_results))

		q1 = db.session.query(
			MemberAccessLog.id,
			MemberAccessLog.memberId,
			MemberAccessLog.username,
			MemberAccessLog.accessTime,
			MemberAccessLog.realIP,
			MemberAccessLog.accessWebsite,
			MemberAccessLog.country,
			MemberAccessLog.province,
			MemberAccessLog.city,
			MemberAccessLog.operator,
		).order_by(MemberAccessLog.accessTime.desc())

		pagination = paginate(q1, criterion, args['page'], args['pageSize'])
		result = []
		for item in pagination.items:
			result.append({
				'id': item.id,
				'memberId': item.memberId,
				'username': item.username,
				'accessTime': item.accessTime,
				'realIP': item.realIP,
				'accessWebsite': item.accessWebsite,
				'country': item.country,
				'province': item.province,
				'city': item.city,
				'operator': item.operator,
			})
		return make_response(result, page=pagination.page, pages=pagination.pages, total=pagination.total)

	def put(self):
		parser = RequestParser(trim=True)
		parser.add_argument('username', type=str, required=True)
		parser.add_argument('sessionKey', type=str, required=True)
		args = parser.parse_args(strict=True)
		criterion = set()
		if args['username']:
			criterion.add(MemberAccessLog.username == args['username'])
		if args['sessionKey']:
			criterion.add(MemberAccessLog.sessionKey == args['sessionKey'])
		m_data = db.session.query(
			MemberAccessLog.id,
			MemberAccessLog.realIP).filter(*criterion).order_by(MemberAccessLog.accessTime.desc()).first()
		result = {}
		if m_data is None:
			return make_response(error_code=400, error_message="数据不存在，无法修改")
		memberAccessLog = MemberAccessLog.query.get(m_data.id)
		urlIP1 = "http://ip.taobao.com/service/getIpInfo.php?ip="+m_data.realIP
		urlIP2 = "http://freeapi.ipip.net/" + m_data.realIP
		try:
			a = 1
			response2 = requests.get(urlIP2)
			data2 = json.loads(response2.text)
			if data2[0] != '本机地址' and data2[0] != '保留地址' and data2[0] != '局域网':
				memberAccessLog.country = data2[0]
				memberAccessLog.province = data2[1]
				memberAccessLog.city = data2[2]
				memberAccessLog.operator = data2[-1]
				try:
					db.session.add(memberAccessLog)
					db.session.commit()
					a = 2
				except:
					db.session.rollback()
					db.session.remove()
		except:
			return make_response([{'success': False}])
		if a == 1:
			try:
				response = requests.get(urlIP1)
				if response:
					data = json.loads(response.text)
					result['province'] = data['data']['region']
					result['city'] = data['data']['city']
					result['operator'] = data['data']['isp']
					result['country'] = data['data']['country']
					if result['country'] != 'XX' and result['country'] != '内网IP':
						memberAccessLog.country = result['country']
					if result['province'] != 'XX' and result['province'] != '内网IP':
						memberAccessLog.province = result['province']
					if result['city'] != 'XX' and result['city'] != '内网IP':
						memberAccessLog.city = result['city']
					if result['operator'] != 'XX' and result['operator'] != '内网IP':
						memberAccessLog.operator = result['operator']
					try:
						db.session.add(memberAccessLog)
						db.session.commit()
					except:
						db.session.rollback()
						db.session.remove()
			except:
				return make_response([{'success': False}])
		return make_response([{'success': True}])


class ExcelMemberAccessLogs(Resource):
	def get(self, id=None):
		parser = RequestParser(trim=True)
		parser.add_argument('page', type=int, default=DEFAULT_PAGE)
		parser.add_argument('pageSize', type=int, default=DEFAULT_PAGE_SIZE)
		parser.add_argument('username', type=str)
		parser.add_argument('accessTimeLower', type=int)
		parser.add_argument('accessTimeUpper', type=int)
		parser.add_argument('realIP', type=str)
		parser.add_argument('realIPLike', type=str)
		parser.add_argument('accessWebsite', type=str)
		parser.add_argument('accessWebsiteLike', type=str)
		parser.add_argument('isLast', type=int)
		args = parser.parse_args(strict=True)

		criterion = set()
		if id:
			criterion.add(MemberAccessLog.id.in_([id]))
		if args['username']:
			criterion.add(MemberAccessLog.username == args['username'])
		if args['accessTimeLower']:
			criterion.add(MemberAccessLog.accessTime >= args['accessTimeLower'])
		if args['accessTimeUpper']:
			criterion.add(MemberAccessLog.accessTime <= args['accessTimeUpper'] + SECONDS_PER_DAY)
		if args['realIP']:
			criterion.add(MemberAccessLog.realIP == args['realIP'])
		if args['realIPLike']:
			criterion.add(MemberAccessLog.realIP.like('%' + args['realIPLike'] + '%'))
		if args['accessWebsite']:
			criterion.add(MemberAccessLog.accessWebsite == args['accessWebsite'])
		if args['accessWebsiteLike']:
			criterion.add(MemberAccessLog.accessWebsite.like('%' + args['accessWebsiteLike'] + '%'))
		if args['isLast'] == 1:
			str_args = '''SELECT * FROM blast_member_session a,
						  (SELECT uid, MAX(loginTime) AS loginTime FROM blast_member_session GROUP BY uid) b
						  WHERE a.uid = b.uid AND a.loginTime = b.loginTime;'''
			uidArray = execute(str_args)[0]
			h_results = []
			for numbers in uidArray:
				h_results.append(numbers[0])
			criterion.add(MemberAccessLog.id.in_(h_results))

		q1 = db.session.query(
			MemberAccessLog.username,
			MemberAccessLog.accessTime,
			MemberAccessLog.realIP,
			MemberAccessLog.accessWebsite,
		).order_by(MemberAccessLog.accessTime.desc()).filter(*criterion).all()
		results =  []
		for item in q1:
			results.append(
				(
					 item[0],
					 changeData_str(item[1]),
					 item[2],
					 item[3]
				 )
			)
		title = ['帐号', '登入时间', 'ip', '网址']
		workbook = Workbook()
		worksheet = workbook.active

		worksheet.append(title)
		worksheet.column_dimensions['A'].width = 20
		worksheet.column_dimensions['B'].width = 20
		worksheet.column_dimensions['C'].width = 18
		worksheet.column_dimensions['D'].width = 20
		for result in results:
			worksheet.append(result)
		filename = '登入记录-' + str(int(time.time())) + '.xlsx'
		workbook.save(os.path.join(current_app.static_folder, filename))
		return make_response([{
			'success': True,
			'resultFilename': filename,
		}])


class MemberLogs(Resource):

	# 订单类型100001，会员充值（银行）
	# 订单类型100002，会员充值（支付）
	# 订单类型200002，会员提现（出款）
	@marshal_with(make_marshal_fields({
		'id': fields.Integer,
		'uid': fields.Integer,
		'auditime': fields.Integer,
		'types': fields.Integer,
		'info': fields.String,
		'makeUser': fields.String,
		'orderId': fields.String,
		'contents': fields.String,
		'ip': fields.String,
		'amount': fields.Integer,
		'username': fields.String,
		'makeUserName': fields.String
	}))
	def get(self):
		historys = RequestParser()
		historys.add_argument('username', type=str)
		m_args = historys.parse_args(strict=True)
		criterion = set()
		if m_args['username']:
			criterion.add(OperationHistory.username == m_args['username'])
		args = OperationHistory().getdata(criterion)
		return make_response(args)


class MemberBatchCreateLogs(Resource):

	@marshal_with(make_marshal_fields({
		'id': fields.Integer,
		'originalFilename': fields.String,
		'status': fields.Integer,
		'timeBegin': fields.Integer,
		'timeEnd': fields.Integer,
		'operator': fields.String,
		'remark': fields.String,
		'resultFilename': fields.String,
	}))
	def get(self):
		parser = RequestParser(trim=True)
		parser.add_argument('page', type=int, default=DEFAULT_PAGE)
		parser.add_argument('pageSize', type=int, default=DEFAULT_PAGE_SIZE)
		args = parser.parse_args(strict=True)
		if not args['pageSize']:
			args['pageSize'] = MemberBatchCreateLog.query.count()
		pagination = paginate(MemberBatchCreateLog.query.order_by(MemberBatchCreateLog.timeBegin.desc()), page=args['page'], per_page=args['pageSize'])
		return make_response_from_pagination(pagination)
	def post(self):
		from werkzeug.datastructures import FileStorage
		from openpyxl import load_workbook
		from datetime import datetime
		parser = RequestParser(trim=True)
		parser.add_argument('original', type=FileStorage, location='files', required=True)
		args = parser.parse_args(strict=True)
		storage = args['original']
		workbook = load_workbook(storage)
# 		fileURL = 'C:\\Users\\lenovo\\Downloads\\ImportMemberCreateTemplate (6).xlsx'
# 		workbook = load_workbook(filename = fileURL)
		worksheet = workbook.active
		worksheet.insert_cols(1)
		title = worksheet.delete_rows(0)
		m_current_user = g.current_user
# 		if m_current_user is None or g.current_user.username:
# 			abort(404, **{'errorCode': 3002,'errorMsg': "没找到管理员",'success': False})
		for row in worksheet.rows:
			member_args = {}
			try:
				if row[5].value:
					pattern = "^[A-Za-z0-9\\u4e00-\\u9fa5][A-Za-z0-9\\u4e00-\\u9fa5-_]{4,13}[A-Za-z0-9\\u4e00-\\u9fa5]$"
					re_result = re.match(pattern, row[5].value)
					if re_result is None:
						row[0].value = '账号必须为6-15位字母和数字组合'
						continue
					member = Member.query.filter(Member.username == row[5].value).first()
					if member:
						row[0].value = '该会员已经存在'
						continue
				if row[4].value:
					parent = Member.query.filter(Member.username == row[4].value,Member.type !=0 ).first()
					if not parent:
						row[0].value = '没有找到代理'
						continue
				else:
					parent = Member.query.filter(Member.id == 1).first()
				#返点
				if row[20].value:
					m_fd = float(row[20].value)
					if m_fd < parent.rebateRate or m_fd == 0:
						member_args['rebateRate'] = m_fd
					else:
						row[0].value = '返点设置不正确'
						continue
				else:
					row[0].value = '返点没设置'
					continue
				member_args['type'] = 0
				member_args['createrUsername'] = g.current_user.username if g.current_user else 'None'
				member_args['username'] = row[5].value
				member_args['password'] = DEFAULT_MEMBER_PASSWORD
				member_args['name'] = row[6].value
				member_args['registrationTime'] = time_to_value()
				member_args['registrationHost'] = host_to_value(request.remote_addr)
				member_args['fundPassword'] = DEFAULT_MEMBER_FUND_PASSWORD
				member_args['parent'] = parent.id
				member_args['parents'] = parent.parents + ',' + str(parent.id)
				member_args['parentsInfo'] = parent.parentsInfo + ',%s' % parent.username
				member_args['levelConfig'] = parent.defaultLevelConfig
				member_args['rebateConfig'] = parent.defaultRebateConfig
				member_args['commissionConfig'] = parent.commissionConfig
				member_args['balance'] = 0
				if row[10].value:
					member_args['balance'] = row[10].value
				member_personal_info_args = {}
				if row[6].value:
					member_personal_info_args['name'] = row[6].value
				if row[7].value:
					member_personal_info_args['phone'] = row[7].value
				if row[8].value:
					member_personal_info_args['email'] = row[8].value
				if row[9].value:
					member_personal_info_args['tencentQQ'] = row[9].value
				if row[12].value:
					if row[12].value == "男":
						member_personal_info_args['gender'] = 1
					else:
						member_personal_info_args['gender'] = 2
				if row[13].value:
					timeArray = time.strptime(row[13].value, "%Y/%m/%d")
					member_personal_info_args['birthdate'] = int(time.mktime(timeArray))
				if row[14].value:
					member_personal_info_args['tencentWeChat'] = row[14].value
				try:
					member = Member(**member_args)
					db.session.add(member)
					db.session.commit()
					try:
						member_personal_info_args['id'] = member.id
						db.session.add(MemberPersonalInfo(**member_personal_info_args))
# 						db.session.commit()
						member_bank_account_args = {}
						if row[15].value:
							bankId = Bank.query.filter(Bank.name == row[15].value).first()
							if bankId:
								member_bank_account_args['bankId'] = bankId.id
							member_bank_account_args['createTime'] = int(time.time())
						if row[16].value:
							member_bank_account_args['province'] = row[16].value
						if row[17].value:
							member_bank_account_args['city'] = row[17].value
						if row[18].value:
							member_bank_account_args['accountNumber'] = row[18].value
						if row[19].value:
							member_bank_account_args['remark'] = row[19].value
						if 'bankId' in member_bank_account_args.keys() and 'accountNumber' in member_bank_account_args.keys():
							member_bank_account_args['memberId'] = member.id
							memberBankDao = MemberBankAccount(**member_bank_account_args)

							member_bank_account_modificationLog = member_bank_account_args
							member_bank_account_modificationLog['userId'] = g.current_user.id
							member_bank_account_modificationLog['time'] = time_to_value()
							del member_bank_account_modificationLog['createTime']
							member_bank_account_modification_log = MemberBankAccountModificationLog(**member_bank_account_modificationLog)
							db.session.add(memberBankDao)
							db.session.add(member_bank_account_modification_log)
# 							db.session.commit()
							row[0].value = '成功'
						else:
							row[0].value = '会员银行信息不标准'
							raise Exception('银行信息错误')
						from app.common.orderUtils import createOrderIdNew
						dml = row[21].value if row[21].value else 0
# 						dml = float(dml)
						deposit = Deposit()
						deposit.status = 2
						deposit.type = 900001
						isAcdemen = 1 if row[11].value == '是' else 0
						deposit.isAcdemen = isAcdemen
						deposit.memberId = member.id
						deposit.username = member.username
						deposit.applicationAmount = float(row[10].value)
						deposit.applicationTime = time_to_value()
						deposit.applicationHost = host_to_value(request.remote_addr)
						deposit.depositAmount = deposit.applicationAmount
						deposit.depositTime = deposit.applicationTime
						deposit.auditUser = g.current_user.id
						deposit.auditTime = time_to_value()
						deposit.auditHost = host_to_value(request.remote_addr)
						deposit.number = createOrderIdNew(uid=member.id)
						db.session.add(deposit)
						from app.models.member_account_change import MemberAccountChangeRecord
						member_account_change_record = MemberAccountChangeRecord()
						member_account_change_record.memberId = member.id
						member_account_change_record.memberBalance = 0
						member_account_change_record.memberFrozenBalance = 0
						member_account_change_record.amount = deposit.depositAmount
						member_account_change_record.accountChangeType = deposit.type
						member_account_change_record.time = deposit.auditTime
						member_account_change_record.host = deposit.auditHost
						member_account_change_record.isAcdemen = deposit.isAcdemen
						member_account_change_record.orderId = deposit.number
						member_account_change_record.rechargeid = deposit.number
						member_account_change_record.auditType = 2
						member_account_change_record.auditCharge = dml
						db.session.add(deposit)
						db.session.add(member_account_change_record)
						db.session.commit()
					except Exception as e:
						db.session.rollback()
						db.session.remove()
						db.session.delete(member)
						db.session.commit()
						row[0].value = '失败'
				except Exception as e:
					db.session.rollback()
					db.session.remove()
					abort(500)
					row[0].value = '失败'
			except Exception as e:
				return make_response(error_code=400)
		filename = '会员汇入-' + str(int(time.time())) + '.xlsx'
		worksheet.insert_rows(title)
		workbook.save(os.path.join(current_app.static_folder, filename))
		try:
			log = MemberBatchCreateLog()
			log.originalFilename = storage.filename
			log.resultFilename = filename
			log.timeBegin = time_to_value()
			log.timeEnd = time_to_value()
			log.operator = g.current_user.username
			db.session.add(log)
			db.session.commit()
		except:
			db.session.rollback()
			db.session.remove()
			abort(500)
		return make_response([]), 201


class GetMembers(Resource):
	def get(self):
		parser = RequestParser(trim=True)
		parser.add_argument('page', type=int, default=DEFAULT_PAGE)
		parser.add_argument('pageSize', type=int, default=DEFAULT_PAGE_SIZE)

		parser.add_argument('username', type=str)
		parser.add_argument('usernameLike', type=str)
		parser.add_argument('registrationTimeLower', type=int)
		parser.add_argument('registrationTimeUpper', type=int)
		parser.add_argument('balanceLower', type=float)
		parser.add_argument('balanceUpper', type=float)
		parser.add_argument('rebateLower', type=float)
		parser.add_argument('rebateUpper', type=float)
		parser.add_argument('rebateConfig', type=int)
		parser.add_argument('rebateRate', type=float)
		parser.add_argument('status', type=int)
		parser.add_argument('levelId', type=str, dest='levelConfig')
		parser.add_argument('parentUsername', type=str)
		parser.add_argument('wallet', type=str)

		parser.add_argument('name', type=str)
		parser.add_argument('nameLike', type=str)
		parser.add_argument('birthdateLower', type=int)
		parser.add_argument('birthdateUpper', type=int)
		parser.add_argument('gender', type=int)
		parser.add_argument('phone', type=str)
		parser.add_argument('phoneLike', type=str)
		parser.add_argument('email', type=str)
		parser.add_argument('emailLike', type=str)
		parser.add_argument('tencentWeChat', type=str)
		parser.add_argument('tencentWeChatLike', type=str)
		parser.add_argument('tencentQQ', type=str)
		parser.add_argument('tencentQQLike', type=str)

		parser.add_argument('bankAccountNumber', type=str)
		parser.add_argument('bankAccountNumberLike', type=str)
		parser.add_argument('bankAccountName', type=str)
		parser.add_argument('bankAccountNameLike', type=str)

		parser.add_argument('lastLoginTimeLower', type=int)
		parser.add_argument('lastLoginTimeUpper', type=int)

		parser.add_argument('lastBetTimeLower', type=int)
		parser.add_argument('lastBetTimeUpper', type=int)

		parser.add_argument('lastBetTimeLowers', type=int)
		parser.add_argument('lastBetTimeUppers', type=int)
		args = parser.parse_args(strict=True)
		criterion = set()
		criterion.add(Member.type == 0)
		criterion.add(Member.isTsetPLay != 1)
		if args['username']:
			criterion.add(Member.username.in_(args['username'].split(',')))
		if args['usernameLike']:
			criterion.add(Member.username.like('%' + args['usernameLike'] + '%'))
		if args['registrationTimeLower']:
			criterion.add(Member.registrationTime >= args['registrationTimeLower'])
		if args['registrationTimeUpper']:
			criterion.add(Member.registrationTime <= args['registrationTimeUpper'] + SECONDS_PER_DAY)
		if args['balanceLower']:
			criterion.add(Member.balance >= args['balanceLower'])
		if args['balanceUpper']:
			criterion.add(Member.balance <= args['balanceUpper'])
		if args['rebateLower']:
			criterion.add(Member.rebate >= args['rebateLower'])
		if args['rebateUpper']:
			criterion.add(Member.rebate <= args['rebateUpper'])
		if args['rebateConfig']:
			criterion.add(Member.rebateConfig == args['rebateConfig'])
		if args['rebateRate'] is not None:
			criterion.add(Member.rebateRate == args['rebateRate'])
		if args['status'] is not None:
			criterion.add(Member.status == args['status'])
		if args['levelConfig']:
			criterion.add(Member.levelConfig.in_(args['levelConfig'].split(',')))
		if args['parentUsername']:
			parent = Member.query.filter(Member.username == args['parentUsername']).first()
			if parent:
				criterion.add(Member.parent == parent.id)
			else:
				return make_response([])
		if args['wallet']:
			pass

		if args['name']:
			criterion.add(MemberPersonalInfo.name == args['name'])
		if args['nameLike']:
			criterion.add(MemberPersonalInfo.name.like('%' + args['nameLike'] + '%'))
		if args['birthdateLower']:
			criterion.add(MemberPersonalInfo.birthdate >= args['birthdateLower'])
		if args['birthdateUpper']:
			criterion.add(MemberPersonalInfo.birthdate <= args['birthdateUpper']  + SECONDS_PER_DAY)
		if args['gender']:
			criterion.add(MemberPersonalInfo.gender == args['gender'])
		if args['phone']:
			criterion.add(MemberPersonalInfo.phone == args['phone'])
		if args['phoneLike']:
			criterion.add(MemberPersonalInfo.phone.like('%' + args['phoneLike'] + '%'))
		if args['email']:
			criterion.add(MemberPersonalInfo.email == args['email'])
		if args['emailLike']:
			criterion.add(MemberPersonalInfo.email.like('%' + args['emailLike'] + '%'))
		if args['tencentWeChat']:
			criterion.add(MemberPersonalInfo.tencentWeChat == args['tencentWeChat'])
		if args['tencentWeChatLike']:
			criterion.add(MemberPersonalInfo.tencentWeChat.like('%' + args['tencentWeChatLike'] + '%'))
		if args['tencentQQ']:
			criterion.add(MemberPersonalInfo.tencentQQ == args['tencentQQ'])
		if args['tencentQQLike']:
			criterion.add(MemberPersonalInfo.tencentQQ.like('%' + args['tencentQQLike'] + '%'))

		if args['bankAccountNumber']:
			str_args = '''select uid from blast_member_bank where account = {}'''.format(args['bankAccountNumber'])
			uidArray = execute(str_args)[0]
			h_results = []
			for numbers in uidArray:
				h_results.append(numbers[0])
			criterion.add(Member.id.in_(h_results))
		if args['bankAccountNumberLike']:
			str_args = '''select uid from blast_member_bank where account like "%{}%"'''.format(
				args["bankAccountNumberLike"])
			uidArray = execute(str_args)[0]
			h_results = []
			for numbers in uidArray:
				h_results.append(numbers[0])
			criterion.add(Member.id.in_(h_results))
		if args['bankAccountName']:
			criterion.add(MemberBankAccount.accountName == args['bankAccountName'])
		if args['bankAccountNameLike']:
			criterion.add(MemberBankAccount.accountName.like('%' + args['bankAccountNameLike'] + '%'))

		if not args['pageSize']:
			args['pageSize'] = Member.query.count()

		criterion_q2 = set()
		if args['lastLoginTimeLower']:
			criterion_q2.add(MemberAccessLog.accessTime >= args['lastLoginTimeLower'])
		if args['lastLoginTimeUpper']:
			criterion_q2.add(MemberAccessLog.accessTime <= args['lastLoginTimeUpper'] +  + SECONDS_PER_DAY)
		if args['lastLoginTimeUpper'] or args['lastLoginTimeLower']:
			q2 = db.session.query(
				MemberAccessLog.memberId.label('uid')
			).filter(*criterion_q2)
			criterion.add(Member.id.in_(q2))

		criterion_q3 = set()
		criterion_q4 = set()
		criterion_ci = set()
		if args['lastBetTimeLower']:
			criterion_q3.add(BlastBets.actionTime >= args['lastBetTimeLower'])
			criterion_q4.add(BlastBetsCredit.betTime >= args['lastBetTimeLower'])
			criterion_ci.add(EntertainmentCityBetsDetail.BetTime >= args['lastBetTimeLower'])
		if args['lastBetTimeUpper']:
			criterion_q3.add(BlastBets.actionTime <= args['lastBetTimeUpper'])
			criterion_q4.add(BlastBetsCredit.betTime <= args['lastBetTimeUpper'])
			criterion_ci.add(EntertainmentCityBetsDetail.BetTime <= args['lastBetTimeUpper'])
		if args['lastBetTimeUpper'] or args['lastBetTimeLower']:
			betsAccount_guan = db.session.query(
				BlastBets.username.label('username')
			).filter(*criterion_q3)
			betsAccount_xinyong = db.session.query(
				BlastBetsCredit.memberUsername.label('username')
			).filter(*criterion_q4)
			betsAccount_city = db.session.query(
				EntertainmentCityBetsDetail.PlayerName.label('username')
			).filter(*criterion_ci)
			result = union_all(betsAccount_guan, betsAccount_xinyong, betsAccount_city)
			user_alias = aliased(result, name='user_alias')
			q3 = db.session.query(
				user_alias.c.username.label('username')
			).distinct()
			criterion.add(Member.username.in_(q3))

		q1 = db.session.query(
			MemberLevel.levelName.label('levelName'),
			MemberLevel.id.label('levelId')
		).subquery()
		q5 = db.session.query(
			func.count(Member.id),
			func.sum(Member.balance),
			func.sum(Member.rebate)
		).order_by(Member.registrationTime.desc())
		q5 = q5.outerjoin(q1, Member.levelConfig == q1.c.levelId)
		q5 = q5.outerjoin(MemberPersonalInfo, Member.id == MemberPersonalInfo.id)
		m_args_list = q5.filter(*criterion).first()
		data = {}
		data['countNum'] = m_args_list[0]
		data['accountNum_one'] = m_args_list[1]
		data['accountNum_two'] = m_args_list[2]
		return make_response(data)



# 监视会员所有钱包
class YlcOutStationBalance(Resource):

	def get(self):
		parser = RequestParser(trim=True)
		parser.add_argument('uid', type=int)
		args = parser.parse_args()
		# 所有娱乐城
		namelist = []
		query1 = db.session.query(EntertainmentCity.code).all()
		for i in query1:
			a = (' '.join(i))
			if a != 'kk':
				namelist.append(a)


		member = Member.query.get(args['uid'])
		if member:
			# 判断是否玩过娱乐城
			query = db.session.query(EntertainmentCityTradeLog.ec).filter(and_(EntertainmentCityTradeLog.username == member.username,EntertainmentCityTradeLog.state == 1)).first()
			if not query:
				arg = []
				result = []
				# 获取娱乐城名称

				for i in query1:
					data = {}
					a = (' '.join(i))
					if a != 'kk':
						data['GameName'] = a
						arg.append(a)
						data['walletnum'] = None
						data['balance'] = None
						data['yhwallet'] = None
						result.append(data)
				return make_response(result)

			# 玩过
			else:
				result1 = []
				ecnames = db.session.query(distinct(EntertainmentCityTradeLog.ec)).filter(and_(
					EntertainmentCityTradeLog.username == member.username,
					EntertainmentCityTradeLog.ec != "MAIN",
					EntertainmentCityTradeLog.state == 1)).all()

				# 玩过那个
				names = []
				for i in ecnames:
					d = (' '.join(i))
					names.append(d)

				walletnum = None
				for i in namelist:
					walletnum = db.session.query(EntertainmentCity.qz).filter(EntertainmentCity.code == i).first()
					walletnum = (' '.join(walletnum))

				# 是否更新过钱包
				jsonlist = db.session.query(YlcOutstationBalance.jsonlist).filter(
					YlcOutstationBalance.username == member.username).first()

				if jsonlist:
					jsonlist = ast.literal_eval(''.join(jsonlist))
					for i in jsonlist:
						if i['GameName'] in names:
							data = {}
							data['GameName'] = i['GameName']
							data['balance'] = i['balance']
							data['walletnum'] = walletnum + member.username
							data['yhwallet'] = None
							result1.append(data)
							names.remove(i['GameName'])
							namelist.remove(i['GameName'])

					for i in names:
						data = {}
						data['GameName'] = i
						data['balance'] = '0.00'
						data['walletnum'] = walletnum + member.username
						data['yhwallet'] = None
						result1.append(data)


					for i in namelist:
						if i not in names:
							data = {}
							data['GameName'] = i
							data['balance'] = None
							data['walletnum'] = None
							data['yhwallet'] = None
							result1.append(data)
				else:
					made = []
					for i in namelist:
						if i in names:
							data = {}
							data2 = {}
							data['GameName'] = i
							data2['GameName'] = i
							data['balance'] = '0.00'
							data2['balance'] = '0.00'
							made.append(data)
							data2['walletnum'] = walletnum + member.username
							data2['yhwallet'] = None
							result1.append(data2)
						else:
							data = {}
							data['GameName'] = i
							data['balance'] = None
							data['walletnum'] = None
							data['yhwallet'] = None
							result1.append(data)

					ylcbalance = YlcOutstationBalance()
					ylcbalance.jsonlist = json.dumps(made)
					ylcbalance.uid = args['uid']
					ylcbalance.username = member.username

					try:
						db.session.add(ylcbalance)
						db.session.commit()
					except:
						db.session.rollback()
						db.session.remove()
				return make_response(result1)

# 更新个人再某个娱乐城的钱包
class UpdateallWallet(Resource):
		def post(self):
			parser = RequestParser()
			parser.add_argument('code', type=str)
			parser.add_argument('uid', type=int)
			args = parser.parse_args()

			result = []
			data = {}
			member = Member.query.get(args['uid'])
			if member:
				user = member.username
				member = UpdataallWallet(user)
				m_data = member.everybalence(args['code'], user)
				m_data = m_data.content.decode()
				m_data = json.loads(m_data)
				# '''{"Guid": "cf2bce64-528b-40dd-8d1a-a9afae96ee5b", "Success": true, "Code": "0", "Message": "Success", "Data": {"balance": "0", "CurrencyCode": null}}'''
				if m_data['Code'] == "0":
					data['balance'] = "%.2f" % float(m_data['Data']['balance'])
					data['GameName'] = args['code']
					result.append(data)

					amount = float("%.2f" % float(m_data['Data']['balance']))
					# 是否更新过
					ylcoutbalance = db.session.query(YlcOutstationBalance).filter(YlcOutstationBalance.username == member.username).first()
					if ylcoutbalance:
						jslist = json.loads(ylcoutbalance.jsonlist)

						if jslist:
							A = [x for x in jslist if (args['code'] == x.get('GameName'))]
							if not A:
								data1 = {}
								data1['balance'] = "%.2f" % float(m_data['Data']['balance'])
								data1['GameName'] = args['code']
								jslist.append(data1)
								ylcoutbalance.jsonlist = json.dumps(jslist)
								db.session.add(ylcoutbalance)
							else:
								A[0]['balance'] = "%.2f" % float(m_data['Data']['balance'])
								ylcoutbalance.jsonlist = json.dumps(jslist)
								db.session.add(ylcoutbalance)


					else:
						ylcbalance = YlcOutstationBalance()
						ylcbalance.jsonlist = json.dumps(result)
						ylcbalance.uid = args['uid']
						ylcbalance.username = member.username
						db.session.add(ylcbalance)
					try:
						db.session.commit()
					except:
						db.session.rollback()
						db.session.remove()
					return make_response(result)

				else:
					return {'errorCode':400,'errorMsg':"由于供应商原因，更新失败"}


# 取回会员再某娱乐城的钱包到主账户
class GetWallet(Resource):
	def post(self):
		parser = RequestParser()
		parser.add_argument('toEC', type=str, required=True)
		parser.add_argument('fromEC', type=str, required=True)
		parser.add_argument('uid', type=int)
		kargs = parser.parse_args()

		mContext = {}
		member = Member.query.get(kargs['uid'])
		mContext["member"] = member
		mContext['loginEC'] = kargs['toEC']
		mContext["ip"] = host_to_value(request.remote_addr)
		mContext["real_ip"] = request.remote_addr
		current_app.logger.info(
			'---------%s从%s取回会员钱包到主账户开始------------' % (member.username, kargs['fromEC']))
		r_data = AmountTransfer.withdrawalToAccount(mContext, kargs)
		current_app.logger.info(
			'---------%s从%s取回会员钱包到主账户结束------------' % (member.username, kargs['fromEC']))

		if r_data['success'] is True:
			ylcoutbalance = db.session.query(YlcOutstationBalance).filter(
				YlcOutstationBalance.username == member.username).first()
			if ylcoutbalance:
				jslist = json.loads(ylcoutbalance.jsonlist)
				if jslist:
					A = [x for x in jslist if (kargs['fromEC'] == x.get('GameName'))]
					if not A:
						data1 = {}
						data1['balance'] = '0.00'
						data1['GameName'] = kargs['fromEC']
						jslist.append(data1)
						ylcoutbalance.jsonlist = json.dumps(jslist)
						db.session.add(ylcoutbalance)
					else:
						A[0]['balance'] = '0.00'
						ylcoutbalance.jsonlist = json.dumps(jslist)
					try:
						db.session.add(ylcoutbalance)
						db.session.commit()
					except:
						db.session.rollback()
						db.session.remove()

			else:
				result = []
				data1 = {}
				data1['balance'] = '0.00'
				data1['GameName'] = kargs['fromEC']
				result.append(data1)
				ylcbalance = YlcOutstationBalance()
				ylcbalance.jsonlist = json.dumps(result)
				ylcbalance.uid = member.id
				ylcbalance.username = member.username
				db.session.add(ylcbalance)

			return r_data

		else:
			return {"errorCode": 3001, "errorMsg": "由于供应商原因，取回失败"}


# 取回用户再所有娱乐城的钱包到主账户
class GetAllEcWallet(Resource):
	def post(self):
		parser = RequestParser()
		parser.add_argument('uid', type=int)
		args = parser.parse_args()

		ecnames = db.session.query(distinct(EntertainmentCityTradeLog.ec)).filter(and_(
					EntertainmentCityTradeLog.uid == args['uid'],
					EntertainmentCityTradeLog.ec != "MAIN",
					EntertainmentCityTradeLog.state == 1)).all()
		if not ecnames:
			return {"errorCode": 400, "errorMsg": "该会员没有创建娱乐城钱包"}

		member = Member.query.get(args['uid'])

		ylcoutbalance = db.session.query(YlcOutstationBalance).filter(
			YlcOutstationBalance.username == member.username).first()

		r_data = None
		jslist = []
		for name in ecnames:
			name = (' '.join(name))
			kargs = {}
			kargs['toEC'] = 'MAIN'
			kargs['fromEC'] = name
			mContext = {}

			mContext["member"] = member
			mContext['loginEC'] = 'MAIN'
			mContext["ip"] = host_to_value(request.remote_addr)
			mContext["real_ip"] = request.remote_addr
			current_app.logger.info(
				'---------%s从%s取回会员钱包到主账户开始------------' % (g.current_user.username, kargs['fromEC']))
			r_data = AmountTransfer.withdrawalToAccount(mContext, kargs)
			current_app.logger.info(
				'---------%s从%s取回会员钱包到主账户结束------------' % (g.current_user.username, kargs['fromEC']))

			if ylcoutbalance:
				data1 = {}
				data1['balance'] = '0.00'
				data1['GameName'] = name
				jslist.append(data1)

			else:
				data2 = {}
				data2['balance'] = '0.00'
				data2['GameName'] = name
				jslist.append(data2)
				ylcoutbalance = YlcOutstationBalance()
				ylcoutbalance.uid = args['uid']
				ylcoutbalance.username = member.username

		try:
			ylcoutbalance.jsonlist = json.dumps(jslist)
			db.session.add(ylcoutbalance)
			db.session.commit()

		except:
			db.session.rollback()
			db.session.remove()

		return make_response(r_data)


# 更新用户再所有娱乐城的钱包
class UpdateAllEcWallet(Resource):
	def post(self):
		parser = RequestParser()
		parser.add_argument('uid', type=int)
		args = parser.parse_args()

		member = Member.query.get(args['uid'])
		if member:
			ecnames = db.session.query(distinct(EntertainmentCityTradeLog.ec)).filter(and_(
					EntertainmentCityTradeLog.uid == args['uid'],
					EntertainmentCityTradeLog.ec != "MAIN",
					EntertainmentCityTradeLog.state == 1)).all()
			if not ecnames:
				return {"errorCode":400,"errorMsg":"该会员没有创建娱乐城钱包"}
			result = []
			for name in ecnames:
				name = (' '.join(name))
				data = {}
				user = member.username
				member = UpdataallWallet(user)
				m_data = member.everybalence(name, user)
				m_data = m_data.content.decode()
				m_data = json.loads(m_data)
				# '''{"Guid": "cf2bce64-528b-40dd-8d1a-a9afae96ee5b", "Success": true, "Code": "0", "Message": "Success", "Data": {"balance": "0", "CurrencyCode": null}}'''
				if m_data['Code'] == "0":
					data['balance'] = float("%.2f" % float(m_data['Data']['balance']))
					data['GameName'] = name
					result.append(data)

			ylcoutbalance = db.session.query(YlcOutstationBalance).filter(
				YlcOutstationBalance.username == member.username).first()
			if ylcoutbalance:
				ylcoutbalance.jsonlist = json.dumps(result)
				try:
					db.session.add(ylcoutbalance)
					db.session.commit()
				except:
					db.session.rollback()
					db.session.remove()
			else:
				ylcoutbalance = YlcOutstationBalance()
				ylcoutbalance.uid = args['uid']
				ylcoutbalance.username = member.username

				try:
					ylcoutbalance.jsonlist = json.dumps(result)
					db.session.add(ylcoutbalance)
					db.session.commit()

				except:
					db.session.rollback()
					db.session.remove()

			return make_response(result)

